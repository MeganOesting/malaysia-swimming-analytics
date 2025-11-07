import os
from openpyxl import load_workbook

HERE = os.path.dirname(__file__)
WB_PATH = os.path.abspath(os.path.join(HERE, '..', '..', 'Malaysia_Times_Database', 'Clubs_By_State.xlsx'))

# Each entry: (club_name, club_code, state_sheet)
ENTRIES = [
    ("Chinese Reacreation Club", "CHREC", "PNG"),
    ("Torpedoes Swimming Club", "TORP", "WPKL"),
    ("Ace Swimming Club", "ACE", "WPKL"),
    ("Alice Smith School", "ALICE", "WPKL"),
    ("SBIS Swimming Team", "SBIS", "WPKL"),
    ("Kuala Lumpur Swimming Club", "KLSC", "WPKL"),
    ("Wave Riders Aqua Centre Alor S", "WRIDE", "KED"),
    ("UCSI Swimming Academy", "UCSI", "WPKL"),
    ("Sailfish Advanced Swim Club", "SAIL", "SEL"),
    ("Tiger Sharks Aquatics", "TIGER", "SBH"),
    ("P.25 Swimming Club", "P25", "JHR"),
    ("Blue One Butterfly", "BLUE", "KED"),
    ("Laksamana Swim Academy", "LAKS", "PHG"),
    ("Mongolia Speed Swim Club", "MONG", "MGL"),
    ("Cyber Titan Sports Academy", "CYBER", "SEL"),
    ("Sarikei Swimming Association", "SSA", "SWK"),
    ("International School of KL", "INTKL", "WPKL"),
    ("Kelab Renang Ara", "ARA", "SEL"),
]

def ensure_row(ws, name, club_code, state_code):
    # Find existing by exact name match in column A
    row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=False), start=1):
        a = row[0].value
        if isinstance(a, str) and a.strip() == name.strip():
            row_idx = i
            break
    if row_idx is None:
        # Append at the bottom
        ws.append([name, club_code, state_code])
    else:
        # Update columns B and C (club_code, state_code) while preserving A
        ws.cell(row=row_idx, column=2).value = club_code
        ws.cell(row=row_idx, column=3).value = state_code

def main():
    if not os.path.exists(WB_PATH):
        raise SystemExit(f"Workbook not found: {WB_PATH}")
    wb = load_workbook(WB_PATH)
    for name, club_code, state in ENTRIES:
        sheet_name = state.strip().upper() if state else None
        if not sheet_name:
            # Skip entries without a state sheet; log by printing
            print(f"Skipping (no state): {name} [{club_code}]")
            continue
        if sheet_name not in wb.sheetnames:
            # Create the sheet if missing
            ws = wb.create_sheet(title=sheet_name)
        else:
            ws = wb[sheet_name]
        ensure_row(ws, name, club_code, sheet_name)
        print(f"Upserted: {name} [{club_code}] in {sheet_name}")
    try:
        wb.save(WB_PATH)
        print("Saved:", WB_PATH)
    except PermissionError:
        alt = WB_PATH.replace('.xlsx', '.updated.xlsx')
        wb.save(alt)
        print("Workbook was locked. Saved updates to:", alt)
        print("Close any app using the original file, then replace it with the .updated.xlsx file.")

if __name__ == "__main__":
    main()
