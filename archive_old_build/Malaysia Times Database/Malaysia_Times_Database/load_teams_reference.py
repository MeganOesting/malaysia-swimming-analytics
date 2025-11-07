import os
import csv
from typing import Dict, List, Tuple, Set

import pandas as pd
from openpyxl import load_workbook, Workbook


BASE_DIR = os.path.dirname(__file__)
CLUBS_XLSX = os.path.join(BASE_DIR, 'Clubs_By_State.xlsx')
ATHLETEINFO_CSV = os.path.join(BASE_DIR, 'AthleteINFO.csv')
TEAMS_REFERENCE_XLSX = os.path.join(BASE_DIR, 'Teams_Reference.xlsx')


STATE_ALIASES: Dict[str, str] = {
    'pulau pinang': 'PNG',
    'penang': 'PNG',
    'wilayah persekutuan': 'WPKL',
    'kuala lumpur': 'WPKL',
    'wp kuala lumpur': 'WPKL',
    'selangor': 'SEL',
    'sarawak': 'SWK',
    'sabah': 'SBH',
    'johor': 'JHR',
    'negeri sembilan': 'NSE',
    'perak': 'PRK',
    'melaka': 'MLK',
    'pahang': 'PHG',
    'kelantan': 'KEL',
    'terengganu': 'TRG',
    'kedah': 'KED',
    'perlis': 'PER',
}


def norm(s: str) -> str:
    return ' '.join(str(s or '').strip().lower().split())


def load_reference(path: str) -> List[Dict[str, str]]:
    df = pd.read_excel(path, engine='openpyxl')
    cols = {norm(c): c for c in df.columns}
    name_col = cols.get('team_name')
    state_col = cols.get('state_code')
    code_col = cols.get('team_code')
    alias_col = cols.get('aliases') if 'aliases' in cols else None
    if not name_col:
        raise ValueError('Teams_Reference.xlsx must include a team_name column')
    out: List[Dict[str, str]] = []
    for _, row in df.iterrows():
        name = str(row.get(name_col) or '').strip()
        if not name:
            continue
        state = str(row.get(state_col) or '').strip().upper() if state_col else ''
        if state and len(state) > 3:
            state = STATE_ALIASES.get(norm(state), state)
        code = str(row.get(code_col) or '').strip().upper() if code_col else ''
        aliases_raw = str(row.get(alias_col) or '').strip() if alias_col else ''
        aliases = [a.strip() for a in aliases_raw.replace(';', ',').split(',') if a.strip()]
        out.append({'team_name': name, 'state_code': state, 'team_code': code, 'aliases': aliases})
    return out


def ensure_clubs_workbook() -> Workbook:
    if os.path.exists(CLUBS_XLSX):
        return load_workbook(CLUBS_XLSX)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def upsert_club(ws, club_name: str, team_code: str) -> None:
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)} if ws.max_row >= 1 else {}
    if not headers:
        ws.append(['club_name', 'club_code'])
        headers = {'club_name': 1, 'club_code': 2}
    name_col = headers.get('club_name', 1)
    code_col = headers.get('club_code', 2)
    # Find existing row by case-insensitive name match
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=name_col).value
        if norm(val) == norm(club_name):
            if team_code:
                ws.cell(row=r, column=code_col).value = team_code
            return
    # Not found; append
    ws.append([club_name, team_code])


def update_clubs_by_state(ref: List[Dict[str, str]]) -> None:
    wb = ensure_clubs_workbook()
    for item in ref:
        state = item.get('state_code', '').upper()
        name = item.get('team_name', '')
        code = item.get('team_code', '')
        aliases = item.get('aliases', [])
        if not state or not name:
            continue
        sheet = state[:31]
        ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(title=sheet)
        upsert_club(ws, name, code)
        for alias in aliases:
            upsert_club(ws, alias, code)
    wb.save(CLUBS_XLSX)


def update_athleteinfo_from_ref(ref: List[Dict[str, str]]) -> Tuple[int, int]:
    if not os.path.exists(ATHLETEINFO_CSV):
        return 0, 0
    # Build quick lookup of team_name and aliases
    name_to_state: Dict[str, str] = {}
    name_to_code: Dict[str, str] = {}
    names_set: Set[str] = set()
    for item in ref:
        state = item.get('state_code', '').upper()
        code = item.get('team_code', '').upper()
        if item.get('team_name'):
            key = norm(item['team_name'])
            names_set.add(key)
            if state:
                name_to_state[key] = state
            if code:
                name_to_code[key] = code
        for alias in item.get('aliases', []) or []:
            k = norm(alias)
            names_set.add(k)
            if state:
                name_to_state[k] = state
            if code:
                name_to_code[k] = code

    with open(ATHLETEINFO_CSV, newline='', encoding='utf-8') as f:
        rows = list(csv.DictReader(f))
    added, updated = 0, 0
    for r in rows:
        tn = norm(r.get('team_name'))
        if tn in names_set:
            st = name_to_state.get(tn)
            cd = name_to_code.get(tn)
            changed = False
            if st and not (r.get('state_code') or '').strip():
                r['state_code'] = st
                changed = True
            if cd and not (r.get('team_code') or '').strip():
                r['team_code'] = cd
                changed = True
            if changed:
                updated += 1
    # Write back
    fieldnames = rows[0].keys() if rows else ['athlete_id','name','birthdate','gender','team_name','team_code','state_code','foreign','source_meet','source_sheet','source_date','aliases','verified','notes']
    with open(ATHLETEINFO_CSV, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)
    return added, updated


def main():
    if not os.path.exists(TEAMS_REFERENCE_XLSX):
        raise SystemExit(f"No Teams_Reference.xlsx found at {TEAMS_REFERENCE_XLSX}")
    ref = load_reference(TEAMS_REFERENCE_XLSX)
    update_clubs_by_state(ref)
    added, updated = update_athleteinfo_from_ref(ref)
    print(f"Applied Teams_Reference: {len(ref)} entries; AthleteINFO updated: {updated} rows")


if __name__ == '__main__':
    main()

