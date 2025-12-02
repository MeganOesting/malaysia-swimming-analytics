"""
Admin API endpoints for Excel file upload and conversion
Using convert_meets_to_sqlite_simple.py for data conversion
"""

import os
import tempfile
import uuid
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, Optional, List
from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Depends, Query, Request
from fastapi.responses import JSONResponse, HTMLResponse, StreamingResponse
from pydantic import BaseModel
import sqlite3
import sys
import logging

# Date validation
from ..utils.date_validator import parse_and_validate_date

# Name matching
from ..utils.name_matcher import match_athlete_by_name
from ..utils.athlete_lookup import find_athlete_ids

# Stroke normalization
from ..utils.stroke_normalizer import normalize_stroke, display_stroke, validate_stroke

logger = logging.getLogger(__name__)

# Add project root to path
# __file__ = src/web/routers/admin.py
# parent = src/web/routers/
# parent.parent = src/web/
# parent.parent.parent = src/
# parent.parent.parent.parent = project root
project_root = Path(__file__).parent.parent.parent.parent
sys.path.append(str(project_root))

# Import conversion logic - using simple converter
from scripts.convert_meets_to_sqlite_simple import (
    process_meet_file_simple,
    get_database_connection,
    insert_data_simple,
    ConversionValidationError,
)
from scripts.convert_clubs_to_sqlite import process_clubs_file, insert_club_data

router = APIRouter()

# Simple password authentication
ADMIN_PASSWORD = "MAS2025"

class AuthRequest(BaseModel):
    password: str

class ConversionResult(BaseModel):
    success: bool
    message: str
    athletes: int = 0
    results: int = 0
    events: int = 0
    meets: int = 0
    unmatched_clubs: List[Dict[str, Any]] = []  # List of unmatched club names from upload (unique names only)
    club_misses: List[Dict[str, Any]] = []  # Full list of club misses with context 
    name_format_mismatches: List[Dict[str, Any]] = []  # List of name format mismatches with athlete_id
    missing_athletes: List[Dict[str, Any]] = []  # List of missing athletes that need to be added to database

class ClubConversionResult(BaseModel):
    success: bool
    message: str
    states: int = 0
    clubs: int = 0
    inserted_states: int = 0
    skipped_states: int = 0
    inserted_clubs: int = 0
    skipped_clubs: int = 0

class SEAGUploadResult(BaseModel):
    success: bool
    message: str
    meet_id: Optional[str] = None
    meet_name: str
    results_inserted: int = 0
    unmatched_athletes: List[Dict[str, Any]] = []  # Athletes not found in database
    errors: List[str] = []  # Any parsing errors

class AliasUpdate(BaseModel):
    alias: str

class CategoryUpdate(BaseModel):
    participant_type: str  # MAST, PARA, OPEN
    scope: str  # I (International), D (Domestic)


class AthleteSearchResponse(BaseModel):
    id: str
    name: str
    gender: Optional[str] = None
    birth_date: Optional[str] = None
    club_name: Optional[str] = None
    state_code: Optional[str] = None
    nation: Optional[str] = None


class AthleteUpdateRequest(BaseModel):
    FULLNAME: Optional[str] = None
    FIRSTNAME: Optional[str] = None
    LASTNAME: Optional[str] = None
    MIDDLEINITIAL: Optional[str] = None
    SUFFIX: Optional[str] = None
    IC: Optional[str] = None
    NATION: Optional[str] = None
    nation: Optional[str] = None  # lowercase alias
    MembEmail: Optional[str] = None
    PreferredName: Optional[str] = None
    Phone: Optional[str] = None
    AcctFirstName: Optional[str] = None
    AcctLastName: Optional[str] = None
    AcctMiddleInitial: Optional[str] = None
    Address: Optional[str] = None
    Address2: Optional[str] = None
    City: Optional[str] = None
    EmergencyContact: Optional[str] = None
    EmergencyPhone: Optional[str] = None
    Guardian1FirstName: Optional[str] = None
    Guardian1HomePhone: Optional[str] = None
    Guardian1LastName: Optional[str] = None
    Guardian1MobilePhone: Optional[str] = None
    Guardian1WorkPhone: Optional[str] = None
    Guardian2FirstName: Optional[str] = None
    Guardian2HomePhone: Optional[str] = None
    Guardian2LastName: Optional[str] = None
    Guardian2MobilePhone: Optional[str] = None
    Guardian2WorkPhone: Optional[str] = None
    AcctIC: Optional[str] = None
    Gender: Optional[str] = None
    ClubCode: Optional[str] = None
    ClubName: Optional[str] = None
    club_code: Optional[str] = None  # lowercase alias
    club_name: Optional[str] = None  # lowercase alias
    state_code: Optional[str] = None  # state code field
    athlete_alias_1: Optional[str] = None
    athlete_alias_2: Optional[str] = None
    passport_number: Optional[str] = None
    shoe_size: Optional[str] = None
    tshirt_size: Optional[str] = None
    tracksuit_size: Optional[str] = None
    cap_name: Optional[str] = None
    School_University_Name: Optional[str] = None
    School_University_Address: Optional[str] = None
    passport_expiry_date: Optional[str] = None
    BIRTHDATE: Optional[str] = None
    postal_code: Optional[str] = None
    address_state: Optional[str] = None

class ClubCreateRequest(BaseModel):
    club_name: str
    club_code: Optional[str] = None
    state_code: Optional[str] = None
    nation: Optional[str] = "MAS"
    alias: Optional[str] = None

class CoachCreateRequest(BaseModel):
    club_name: str
    name: str
    role: str  # 'head_coach', 'assistant_coach', 'manager'
    email: Optional[str] = None
    whatsapp: Optional[str] = None
    passport_photo: Optional[str] = None
    passport_number: Optional[str] = None
    ic: Optional[str] = None
    shoe_size: Optional[str] = None
    tshirt_size: Optional[str] = None
    tracksuit_size: Optional[str] = None
    course_level_1_sport_specific: bool = False
    course_level_2: bool = False
    course_level_3: bool = False
    course_level_1_isn: bool = False
    course_level_2_isn: bool = False
    course_level_3_isn: bool = False
    seminar_oct_2024: bool = False
    other_courses: Optional[str] = None
    state_coach: bool = False
    logbook_file: Optional[str] = None


class ManualResult(BaseModel):
    athlete_id: str
    distance: int
    stroke: str
    event_gender: str
    time_string: str
    comp_place: Optional[int] = None


class ManualResultsSubmission(BaseModel):
    meet_name: str
    meet_date: str
    meetcity: str
    meet_course: str
    meet_alias: str
    results: List[ManualResult]


@router.get("/admin/test")
async def test_admin():
    """Test admin endpoint"""
    return {"message": "Admin router is working"}

@router.options("/admin/authenticate")
async def options_authenticate():
    """Handle CORS preflight for authenticate endpoint"""
    return {"ok": True}

@router.post("/admin/authenticate")
async def authenticate(request: AuthRequest):
    """Authenticate admin user"""
    if request.password == ADMIN_PASSWORD:
        return {"success": True, "message": "Authentication successful"}
    else:
        raise HTTPException(status_code=401, detail="Invalid password")

@router.post("/admin/convert-excel", response_model=ConversionResult)
async def convert_excel(
    file: UploadFile = File(...),
    meet_name: str = Form(None),
    meet_code: str = Form(None),
    existing_meet_id: str = Form(None)
):
    """Convert uploaded Excel file to database entries using fixed conversion script"""
    print(f"\n[UPLOAD REQUEST] Received: {file.filename}")
    
    # Validate file type
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")
    
    # Save uploaded file temporarily (stream in chunks to reduce startup delay and memory)
    print(f"[UPLOAD] Saving file to temporary location...")
    with tempfile.NamedTemporaryFile(delete=False, suffix=Path(file.filename).suffix) as temp_file:
        temp_file_path = temp_file.name
        # Stream read in 1MB chunks
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            temp_file.write(chunk)
    print(f"[UPLOAD] File saved, starting processing...")
    
    try:
        # Process the file
        result = await process_uploaded_file(temp_file_path, file.filename, meet_name, meet_code, existing_meet_id)
        # Ensure a concrete JSON body is returned (avoid empty preview)
        try:
            body = result.dict() if hasattr(result, 'dict') else dict(result)
        except Exception:
            body = {
                "success": getattr(result, 'success', True),
                "message": getattr(result, 'message', "Conversion complete"),
                "athletes": getattr(result, 'athletes', 0),
                "results": getattr(result, 'results', 0),
                "events": getattr(result, 'events', 0),
                "meets": getattr(result, 'meets', 1),
            }
        return JSONResponse(content=body)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Conversion failed: {str(e)}")
    
    finally:
        # Clean up temporary file
        try:
            os.unlink(temp_file_path)
        except:
            pass

@router.post("/admin/upload-seag", response_model=SEAGUploadResult)
async def upload_seag(file: UploadFile = File(...), 

    meet_name: str = Form(...),
    meetcity: str = Form(...),
    meet_month: str = Form(...),
    meet_day: str = Form(...),
   year: str = Form(...)):
    """Upload SEAG results from Excel file - matches athletes by name"""
    print(f"\n[SEAG UPLOAD] Received: {file.filename}")

    # Validate file type
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")

    # Save uploaded file temporarily
    print(f"[SEAG UPLOAD] Saving file to temporary location...")
    with tempfile.NamedTemporaryFile(delete=False, suffix=Path(file.filename).suffix) as temp_file:
        temp_file_path = temp_file.name
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            temp_file.write(chunk)

    print(f"[SEAG UPLOAD] File saved, starting processing...")

    try:
        import pandas as pd
        from src.web.utils.calculation_utils import parse_time_to_seconds, calculate_aqua_points

        # Read Excel file
        try:
            df = pd.read_excel(temp_file_path, sheet_name="Sheet", skiprows=[0], header=0)
        except Exception as e:
            return SEAGUploadResult(
                success=False,
                message=f"Failed to read Excel file: {str(e)}",
                results_inserted=0,
                unmatched_athletes=[],
                errors=[str(e)]
            )

        print(f"[SEAG UPLOAD] Loaded {len(df)} rows from Excel")

        # Get database connection
        conn = get_database_connection()
        cursor = conn.cursor()

        # Create or get SEAG meet
        meet_id = f"SEAG_{year}"
        result_meet_date = f"{year}-{meet_month.zfill(2)}-{meet_day.zfill(2)}T00:00:00Z"

        # Ensure meet exists
        cursor.execute("SELECT id FROM meets WHERE id = ?", (meet_id,))
        if not cursor.fetchone():
            cursor.execute("""
                INSERT INTO meets (id, meet_name, meet_type, meet_date, location, meet_city)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (meet_id, meet_name, "International", result_meet_date, meetcity, meetcity))
            conn.commit()
            print(f"[SEAG UPLOAD] Created meet: {meet_id}")
        else:
            print(f"[SEAG UPLOAD] Meet already exists: {meet_id}")

        # Process results
        results_inserted = 0
        unmatched_athletes = []
        errors = []

        for idx, row in df.iterrows():
            try:
                # Extract data
                gender = str(row.get('GENDER', '')).strip().upper()[0] if pd.notna(row.get('GENDER')) else None
                distance = int(row.get('DISTANCE', 0)) if pd.notna(row.get('DISTANCE')) else None
                stroke_raw = str(row.get('STROKE', '')).strip() if pd.notna(row.get('STROKE')) else None
                fullname = str(row.get('FULLNAME', '')).strip() if pd.notna(row.get('FULLNAME')) else None
                time_str = str(row.get('SWIMTIME', '')).strip() if pd.notna(row.get('SWIMTIME')) else None
                place = int(row.get('PLACE', 0)) if pd.notna(row.get('PLACE')) else None
                # Read STATUS column (OK, DQ, DNS, DNF, SCR) - defaults to OK if not present
                status_raw = str(row.get('STATUS', '')).strip().upper() if pd.notna(row.get('STATUS')) else None
                result_status = status_raw if status_raw in ('DQ', 'DNS', 'DNF', 'SCR') else 'OK'
                # Extract rudolph points from file (if present)
                rudolph_points = int(row.get('PTS_RUDOLPH', 0)) if pd.notna(row.get('PTS_RUDOLPH')) else None
                # Accept meet_course (standard) or COURSE (SwimRankings legacy)
                raw_course = row.get('meet_course') if pd.notna(row.get('meet_course')) else row.get('COURSE')
                meet_course = str(raw_course).strip().upper() if pd.notna(raw_course) else 'LCM'
                # Accept club_name (standard) or CLUBNAME (legacy)
                raw_club_name = row.get('club_name') if pd.notna(row.get('club_name')) else row.get('CLUBNAME')
                club_name = str(raw_club_name).strip() if pd.notna(raw_club_name) else None
                # Accept club_code (standard) or CLUBCODE (legacy)
                raw_club_code = row.get('club_code') if pd.notna(row.get('club_code')) else row.get('CLUBCODE')
                club_code = str(raw_club_code).strip() if pd.notna(raw_club_code) else None
                # Accept nation (standard) or NATION (legacy)
                raw_nation = row.get('nation') if pd.notna(row.get('nation')) else row.get('NATION')
                nation = str(raw_nation).strip() if pd.notna(raw_nation) else None

                # Validate required fields (time not required for DQ/DNS/DNF/SCR)
                if not all([gender, distance, stroke_raw, fullname]):
                    missing = []
                    if not gender: missing.append('GENDER')
                    if not distance: missing.append('DISTANCE')
                    if not stroke_raw: missing.append('STROKE')
                    if not fullname: missing.append('FULLNAME')
                    error_msg = f"Row {idx+2}: Missing fields: {', '.join(missing)}"
                    print(f"[SEAG UPLOAD] {error_msg}")
                    errors.append(error_msg)
                    continue

                # For DQ/DNS/DNF/SCR, clear time and place
                if result_status in ('DQ', 'DNS', 'DNF', 'SCR'):
                    time_str = None
                    place = None
                elif not time_str:
                    # Time required for OK results
                    error_msg = f"Row {idx+2}: Missing SWIMTIME for OK result"
                    print(f"[SEAG UPLOAD] {error_msg}")
                    errors.append(error_msg)
                    continue

                # Normalize stroke using global normalizer
                stroke_name = normalize_stroke(stroke_raw)

                # Look up event first (needed for unmatched alert)
                cursor.execute("""
                    SELECT id FROM events
                    WHERE event_distance = ? AND event_stroke = ? AND gender = ?
                    LIMIT 1
                """, (distance, stroke_name, gender))
                event_row = cursor.fetchone()
                event_id = event_row[0] if event_row else None
                event_desc = event_id or f"{distance} {stroke_name} {gender}"

                # Look up athlete using flexible name matching
                athlete_id = match_athlete_by_name(conn, fullname, gender)

                if not athlete_id:
                    # Track unmatched with FULLNAME and EVENT
                    unmatched_athletes.append({
                        'fullname': fullname,
                        'event': event_desc,
                        'row': idx + 2
                    })
                    continue

                # Create event if it doesn't exist
                if not event_id:
                    event_id = str(uuid.uuid4())
                    cursor.execute("""
                        INSERT INTO events (id, event_distance, event_stroke, gender, event_course)
                        VALUES (?, ?, ?, ?, ?)
                    """, (event_id, distance, stroke_name, gender, meet_course))

                # Convert time to seconds using shared utility (skip for DQ/DNS/DNF/SCR)
                time_seconds = None
                if time_str and result_status == 'OK':
                    time_seconds = parse_time_to_seconds(time_str)
                    if not time_seconds:
                        error_msg = f"Row {idx+2}: Invalid time format: {time_str}"
                        errors.append(error_msg)
                        continue

                # Calculate AQUA points (DO NOT read from file for SEAG)
                aqua_points = calculate_aqua_points(conn, gender, distance, stroke_name, time_seconds)

                # Get athlete details for age calculation
                year_age = None
                day_age = None
                team_name = club_name
                team_code = club_code
                team_state_code = None
                team_nation = nation

                cursor.execute("""
                    SELECT club_name, state_code, nation, BIRTHDATE
                    FROM athletes WHERE id = ?
                """, (athlete_id,))
                athlete_data = cursor.fetchone()
                if athlete_data:
                    team_name = athlete_data[0] or club_name
                    team_state_code = athlete_data[1]
                    team_nation = athlete_data[2] or nation
                    athlete_birthdate = athlete_data[3]

                    # Calculate ages from athlete birthdate
                    if athlete_birthdate:
                        try:
                            from dateutil import parser as date_parser
                            birth_dt = date_parser.parse(str(athlete_birthdate))
                            # Make birth_dt naive (remove timezone) for comparison
                            if birth_dt.tzinfo is not None:
                                birth_dt = birth_dt.replace(tzinfo=None)
                            meet_dt = datetime(int(year), int(meet_month), int(meet_day))
                            year_age = meet_dt.year - birth_dt.year
                            if (meet_dt.month, meet_dt.day) < (birth_dt.month, birth_dt.day):
                                year_age -= 1
                            day_age = (meet_dt - birth_dt).days
                        except Exception:
                            pass

                # Check for duplicate before inserting
                # For DQ/DNS/DNF/SCR: check athlete, event, meet only
                # For OK results: also check time_seconds
                if result_status in ('DQ', 'DNS', 'DNF', 'SCR'):
                    cursor.execute("""
                        SELECT id FROM results
                        WHERE athlete_id = ? AND event_id = ? AND meet_id = ? AND result_status = ?
                        LIMIT 1
                    """, (athlete_id, event_id, meet_id, result_status))
                else:
                    cursor.execute("""
                        SELECT id FROM results
                        WHERE athlete_id = ? AND event_id = ? AND meet_id = ? AND time_seconds = ?
                        LIMIT 1
                    """, (athlete_id, event_id, meet_id, time_seconds))
                existing = cursor.fetchone()
                if existing:
                    # Skip duplicate - already in database
                    continue

                # Insert result with all fields (matching preview columns exactly)
                result_id = str(uuid.uuid4())
                cursor.execute("""
                    INSERT INTO results (
                        id, meet_id, athlete_id, event_id, time_seconds, time_string,
                        aqua_points, rudolph_points, meet_course, meet_date,
                        day_age, year_age, club_name, club_code, state_code,
                        nation, is_relay, comp_place, meet_name, meet_city, result_status
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (result_id, meet_id, athlete_id, event_id, time_seconds, time_str,
                      aqua_points, rudolph_points, meet_course, result_meet_date,
                      day_age, year_age, team_name, team_code, team_state_code,
                      team_nation, 0, place, meet_name, meetcity, result_status))

                results_inserted += 1

            except Exception as e:
                error_msg = f"Row {idx+2}: {str(e)}"
                print(f"[SEAG UPLOAD ERROR] {error_msg}")
                errors.append(error_msg)
                import traceback
                traceback.print_exc()
                continue

        conn.commit()
        conn.close()

        # Build detailed message with all error counts
        message_parts = [f"Results loaded: {results_inserted}"]
        if unmatched_athletes:
            message_parts.append(f"Unmatched athletes: {len(unmatched_athletes)}")
        if errors:
            message_parts.append(f"Errors: {len(errors)}")

        message = "SEAG upload complete - " + " | ".join(message_parts) if message_parts else "SEAG upload complete - No results inserted"

        # Print first few errors for debugging
        if errors:
            print(f"\n[SEAG UPLOAD] First 5 errors:")
            for err in errors[:5]:
                print(f"  {err}")

        return SEAGUploadResult(
            success=True,
            message=message,
            meet_id=meet_id,
            meet_name=meet_name,
            results_inserted=results_inserted,
            unmatched_athletes=unmatched_athletes,
            errors=errors
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"SEAG upload failed: {str(e)}")

    finally:
        # Clean up temporary file
        try:
            os.unlink(temp_file_path)
        except:
            pass


@router.post("/admin/preview-seag-upload")
async def preview_seag_upload(
    file: UploadFile = File(...),
    meet_name: str = Form(...),
    meetcity: str = Form(...),
    meet_month: str = Form(...),
    meet_day: str = Form(...),
    year: str = Form(...)):
    """Generate preview Excel file showing what data would be uploaded"""
    print(f"\n[PREVIEW SEAG] Received: {file.filename}")

    # Validate file type
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")

    # Save uploaded file temporarily
    print(f"[PREVIEW SEAG] Saving file to temporary location...")
    with tempfile.NamedTemporaryFile(delete=False, suffix=Path(file.filename).suffix) as temp_file:
        temp_file_path = temp_file.name
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            temp_file.write(chunk)

    print(f"[PREVIEW SEAG] File saved, starting processing...")

    try:
        import pandas as pd
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import io
        from datetime import datetime

        # Read Excel file
        try:
            df = pd.read_excel(temp_file_path, sheet_name="Sheet", skiprows=[0], header=0)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")

        print(f"[PREVIEW SEAG] Loaded {len(df)} rows from Excel")

        # Import time parser and AQUA points calculator
        from src.web.utils.calculation_utils import parse_time_to_seconds, calculate_aqua_points

        # Get database connection
        conn = get_database_connection()
        cursor = conn.cursor()

        # Build meet date from user input
        result_meet_date = f"{year}-{meet_month.zfill(2)}-{meet_day.zfill(2)}T00:00:00Z"

        # Process rows to collect preview data matching results table format
        preview_rows = []
        unmatched_athletes = []  # Track unmatched athletes for alert
        row_num = 0

        # Log actual columns for debugging
        print(f"[PREVIEW SEAG] Excel columns: {list(df.columns)}")

        for idx, row in df.iterrows():
            try:
                # Extract data from standard SwimRankings/SEAG format columns
                gender = str(row.get('GENDER', '')).strip().upper()[0] if pd.notna(row.get('GENDER')) else None
                distance = int(row.get('DISTANCE', 0)) if pd.notna(row.get('DISTANCE')) else None
                stroke_raw = str(row.get('STROKE', '')).strip() if pd.notna(row.get('STROKE')) else None
                fullname = str(row.get('FULLNAME', '')).strip() if pd.notna(row.get('FULLNAME')) else None
                time_str = str(row.get('SWIMTIME', '')).strip() if pd.notna(row.get('SWIMTIME')) else None
                # DO NOT read PTS_FINA from file for SEAG - calculate it instead
                rudolph_points = int(row.get('PTS_RUDOLPH', 0)) if pd.notna(row.get('PTS_RUDOLPH')) else None
                place = int(row.get('PLACE', 0)) if pd.notna(row.get('PLACE')) else None
                # Accept meet_course (standard) or COURSE (SwimRankings legacy)
                raw_course = row.get('meet_course') if pd.notna(row.get('meet_course')) else row.get('COURSE')
                meet_course = str(raw_course).strip().upper() if pd.notna(raw_course) else 'LCM'
                # Accept club_name (standard) or CLUBNAME (legacy)
                raw_club_name = row.get('club_name') if pd.notna(row.get('club_name')) else row.get('CLUBNAME')
                club_name = str(raw_club_name).strip() if pd.notna(raw_club_name) else None
                # Accept club_code (standard) or CLUBCODE (legacy)
                raw_club_code = row.get('club_code') if pd.notna(row.get('club_code')) else row.get('CLUBCODE')
                club_code = str(raw_club_code).strip() if pd.notna(raw_club_code) else None
                # Accept nation (standard) or NATION (legacy)
                raw_nation = row.get('nation') if pd.notna(row.get('nation')) else row.get('NATION')
                nation = str(raw_nation).strip() if pd.notna(raw_nation) else None

                # Skip rows missing required fields
                if not all([gender, distance, stroke_raw, fullname, time_str]):
                    if idx == 0:
                        print(f"[PREVIEW SEAG] First row missing data: gender={gender}, distance={distance}, stroke={stroke_raw}, fullname={fullname}, time={time_str}")
                    continue

                row_num += 1

                # Normalize stroke using global normalizer
                stroke_name = normalize_stroke(stroke_raw)

                # Look up athlete
                athlete_id = match_athlete_by_name(conn, fullname, gender)

                # Look up event
                cursor.execute("""
                    SELECT id FROM events
                    WHERE event_distance = ? AND event_stroke = ? AND gender = ?
                    LIMIT 1
                """, (distance, stroke_name, gender))
                event_result = cursor.fetchone()
                event_id = event_result[0] if event_result else None

                # Build event description for unmatched alert
                event_desc = event_id or f"{distance} {stroke_name} {gender}"

                # Track unmatched athletes for alert
                if not athlete_id:
                    unmatched_athletes.append({
                        'fullname': fullname,
                        'distance': distance,
                        'stroke': stroke_name,
                        'gender': gender,
                        'event': event_desc,
                        'row': idx + 2  # Excel row number (1-indexed + header)
                    })

                # Get athlete details if matched (age calc only works when matched)
                team_name = club_name
                team_code = club_code
                team_state_code = None
                team_nation = nation
                year_age = None
                day_age = None
                athlete_birthdate = None

                if athlete_id:
                    cursor.execute("""
                        SELECT club_name, state_code, nation, BIRTHDATE
                        FROM athletes WHERE id = ?
                    """, (athlete_id,))
                    athlete_data = cursor.fetchone()
                    if athlete_data:
                        team_name = athlete_data[0] or club_name
                        team_state_code = athlete_data[1]
                        team_nation = athlete_data[2] or nation
                        athlete_birthdate = athlete_data[3]

                        # Calculate year_age and day_age from athlete birthdate and meet date
                        if athlete_birthdate:
                            try:
                                from dateutil import parser as date_parser
                                birth_dt = date_parser.parse(str(athlete_birthdate))
                                # Make birth_dt naive (remove timezone) for comparison
                                if birth_dt.tzinfo is not None:
                                    birth_dt = birth_dt.replace(tzinfo=None)
                                meet_year = int(year)
                                meet_month_int = int(meet_month)
                                meet_day_int = int(meet_day)
                                meet_dt = datetime(meet_year, meet_month_int, meet_day_int)
                                # Calculate age in years
                                year_age = meet_dt.year - birth_dt.year
                                if (meet_dt.month, meet_dt.day) < (birth_dt.month, birth_dt.day):
                                    year_age -= 1
                                # Calculate day_age (days between birthdate and meet date)
                                day_age = (meet_dt - birth_dt).days
                            except Exception as age_err:
                                print(f"[PREVIEW SEAG] Could not calculate age for {fullname}: {age_err}")

                # Calculate time in seconds
                time_seconds = parse_time_to_seconds(time_str)

                # Calculate AQUA points using the calculator (DO NOT read from file for SEAG)
                aqua_points = None
                if time_seconds and time_seconds > 0 and stroke_name and distance and gender:
                    aqua_points = calculate_aqua_points(conn, gender, distance, stroke_name, time_seconds)

                # Generate preview row ID
                preview_id = f"PREVIEW_{row_num}"

                # Build row matching results table columns exactly
                preview_rows.append({
                    'id': preview_id,
                    'meet_id': f"SEAG_{year}",
                    'athlete_id': athlete_id or f"UNMATCHED: {fullname}",
                    'event_id': event_id or f"MISSING: {distance} {stroke_name} {gender}",
                    'time_seconds': time_seconds,
                    'time_string': time_str,
                    'aqua_points': aqua_points,
                    'rudolph_points': rudolph_points,
                    'meet_course': meet_course,
                    'meet_date': result_meet_date,
                    'day_age': day_age,
                    'year_age': year_age,
                    'club_name': team_name,
                    'club_code': team_code,
                    'state_code': team_state_code,
                    'nation': team_nation,
                    'is_relay': 0,
                    'comp_place': place,
                    'meet_name': meet_name,
                    'meet_city': meetcity,
                })

            except Exception as e:
                import traceback
                print(f"[PREVIEW SEAG] Row {idx+2} error: {str(e)}")
                traceback.print_exc()

        # Print match summary
        total_rows = len(preview_rows)
        matched_count = total_rows - len(unmatched_athletes)
        unmatched_count = len(unmatched_athletes)

        print(f"\n{'='*60}")
        print(f"[SEAG PREVIEW] MATCH SUMMARY")
        print(f"{'='*60}")
        print(f"Total athletes processed: {total_rows}")
        print(f"Matched athletes: {matched_count}")
        print(f"Unmatched athletes: {unmatched_count}")

        if unmatched_athletes:
            print(f"\n[SEAG PREVIEW] UNMATCHED ATHLETES:")
            print(f"{'-'*60}")
            for u in unmatched_athletes:
                print(f"  - {u['fullname']} | {u['distance']}m {u['stroke']} ({u['gender']})")

        print(f"{'='*60}\n")

        # Create Excel workbook matching results export format
        wb = Workbook()
        ws = wb.active
        ws.title = "Results Data"

        # Headers matching results table columns exactly
        headers = [
            'id', 'meet_id', 'athlete_id', 'event_id', 'time_seconds', 'time_string',
            'aqua_points', 'rudolph_points', 'meet_course',
            'meet_date', 'day_age', 'year_age', 'club_name',
            'club_code', 'state_code', 'nation', 'is_relay',
            'comp_place', 'meet_name', 'meet_city'
        ]

        # Write headers with red styling (matching results export)
        ws.append(headers)
        header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Write data rows
        for preview_row in preview_rows:
            row_data = [preview_row.get(h, '') for h in headers]
            ws.append(row_data)

        # Auto-adjust column widths
        for i, column in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value or '')) for cell in column)
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(max_length + 2, 50)

        # Add Unmatched Athletes alert sheet if any unmatched
        if unmatched_athletes:
            ws_unmatched = wb.create_sheet(title="UNMATCHED ATHLETES")

            # Headers for unmatched sheet
            unmatched_headers = ['FULLNAME', 'EVENT', 'EXCEL ROW']
            ws_unmatched.append(unmatched_headers)

            # Style headers with warning color (orange/red)
            warning_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
            for cell in ws_unmatched[1]:
                cell.fill = warning_fill
                cell.font = header_font

            # Write unmatched athlete data
            for unmatched in unmatched_athletes:
                ws_unmatched.append([
                    unmatched.get('fullname', ''),
                    unmatched.get('event', ''),
                    unmatched.get('row', '')
                ])

            # Auto-adjust column widths for unmatched sheet
            for i, column in enumerate(ws_unmatched.columns, 1):
                max_length = max(len(str(cell.value or '')) for cell in column)
                ws_unmatched.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(max_length + 2, 50)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Return as Excel file download
        from fastapi.responses import StreamingResponse
        filename = f"seag_{year}_preview.xlsx"
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Preview generation failed: {str(e)}")

    finally:
        # Clean up temporary file
        try:
            os.unlink(temp_file_path)
        except:
            pass


@router.post("/admin/preview-swimrankings-upload")
async def preview_swimrankings_upload(file: UploadFile = File(...)):
    """Generate preview Excel file for SwimRankings upload.

    CRITICAL: This uses the EXACT SAME processing logic as upload (process_meet_file_simple).
    The only difference is output destination: Excel file instead of database.
    """
    print(f"\n[PREVIEW SWIMRANKINGS] Received: {file.filename}")

    # Validate file type
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")

    # Save uploaded file temporarily
    print(f"[PREVIEW SWIMRANKINGS] Saving file to temporary location...")
    with tempfile.NamedTemporaryFile(delete=False, suffix=Path(file.filename).suffix) as temp_file:
        temp_file_path = temp_file.name
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            temp_file.write(chunk)

    print(f"[PREVIEW SWIMRANKINGS] File saved, starting processing...")

    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        import io

        # Create meet_info dict - SAME as upload does
        meet_info = {
            'id': 'PREVIEW',
            'name': 'Preview',
            'meet_type': None,
            'meet_date': None,
            'location': None,
        }

        # Call process_meet_file_simple - EXACT SAME as upload
        # This handles: sheet skipping (4x, TOP, 5000m), athlete matching, event lookup, etc.
        file_path_obj = Path(temp_file_path)
        athletes, results, events, collector = process_meet_file_simple(file_path_obj, meet_info)

        print(f"[PREVIEW] Processed: {len(results)} results, {len(collector.missing_athletes)} missing athletes")

        # Count matched vs unmatched
        # Results only contains MATCHED rows (process_meet_file_simple skips unmatched)
        matched_count = len(results)
        unmatched_count = len(collector.missing_athletes)
        total_rows = matched_count + unmatched_count

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Results Data"

        # Headers matching results table columns exactly
        headers = [
            'id', 'meet_id', 'athlete_id', 'foreign_athlete_id', 'event_id', 'time_seconds', 'time_string',
            'aqua_points', 'rudolph_points', 'meet_course',
            'meet_date', 'day_age', 'year_age', 'club_name',
            'club_code', 'state_code', 'nation', 'is_relay',
            'comp_place', 'meet_name', 'meet_city'
        ]

        # Write headers with red styling
        ws.append(headers)
        header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Write data rows from results (output from process_meet_file_simple)
        for result in results:
            row_data = [
                result.get('id', ''),
                result.get('meet_id', ''),
                result.get('athlete_id', ''),
                result.get('foreign_athlete_id', ''),
                result.get('event_id', ''),
                result.get('time_seconds', ''),
                result.get('time_string', ''),
                result.get('aqua_points', ''),
                result.get('rudolph_points', ''),
                result.get('course', ''),
                result.get('meet_date', result.get('result_meet_date', '')),
                result.get('day_age', ''),
                result.get('year_age', ''),
                result.get('club_name', ''),
                result.get('club_code', ''),
                result.get('state_code', ''),
                result.get('nation', ''),
                result.get('is_relay', 0),
                result.get('place', result.get('comp_place', '')),
                result.get('meet_name', ''),
                result.get('meet_city', ''),
            ]
            ws.append(row_data)

        # Auto-adjust column widths
        for i, column in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value or '')) for cell in column)
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(max_length + 2, 50)

        # Calculate MAS vs FOREIGN unmatched counts for summary
        mas_unmatched_count = len([m for m in collector.missing_athletes if not m.get('likely_foreign', False)])
        foreign_unmatched_count = len([m for m in collector.missing_athletes if m.get('likely_foreign', False)])

        # Add SUMMARY sheet first
        ws_summary = wb.create_sheet(title="SUMMARY", index=0)
        summary_data = [
            ['PREVIEW SUMMARY', ''],
            ['', ''],
            ['Total Results Found:', total_rows],
            ['Results MATCHED (will upload):', matched_count],
            ['Results UNMATCHED (will NOT upload):', unmatched_count],
            ['', ''],
            ['  - MAS unmatched (add to athletes):', mas_unmatched_count],
            ['  - FOREIGN unmatched (add to foreign_athletes):', foreign_unmatched_count],
            ['', ''],
            ['NOTE: See "UNMATCHED - MAS" and "UNMATCHED - FOREIGN" sheets'],
        ]
        for row in summary_data:
            ws_summary.append(row)

        # Style summary headers
        summary_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        ws_summary['A1'].fill = summary_fill
        ws_summary['A1'].font = Font(bold=True, color="FFFFFF", size=14)
        ws_summary.column_dimensions['A'].width = 45
        ws_summary.column_dimensions['B'].width = 20

        # Split unmatched athletes into MAS and FOREIGN sheets
        if collector.missing_athletes:
            # Separate MAS vs FOREIGN based on likely_foreign flag
            mas_unmatched = [m for m in collector.missing_athletes if not m.get('likely_foreign', False)]
            foreign_unmatched = [m for m in collector.missing_athletes if m.get('likely_foreign', False)]

            unmatched_headers = ['SHEET', 'ROW', 'FULLNAME', 'BIRTHDATE', 'GENDER', 'EXCEL_NATION', 'CLUB_NAME', 'MEET_NAME']

            # MAS UNMATCHED sheet (orange headers - need to add to athletes table)
            if mas_unmatched:
                ws_mas = wb.create_sheet(title="UNMATCHED - MAS")
                ws_mas.append(unmatched_headers)
                mas_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")  # Orange
                for cell in ws_mas[1]:
                    cell.fill = mas_fill
                    cell.font = header_font
                for missing in mas_unmatched:
                    ws_mas.append([
                        missing.get('sheet', ''),
                        missing.get('row', ''),
                        missing.get('full_name', ''),
                        missing.get('birthdate', ''),
                        missing.get('gender', ''),
                        missing.get('excel_nation', ''),
                        missing.get('club_name', ''),
                        missing.get('meet_name', ''),
                    ])
                # Column widths
                ws_mas.column_dimensions['A'].width = 15  # SHEET
                ws_mas.column_dimensions['B'].width = 6   # ROW
                ws_mas.column_dimensions['C'].width = 35  # FULLNAME
                ws_mas.column_dimensions['D'].width = 12  # BIRTHDATE
                ws_mas.column_dimensions['E'].width = 8   # GENDER
                ws_mas.column_dimensions['F'].width = 10  # EXCEL_NATION
                ws_mas.column_dimensions['G'].width = 30  # CLUB_NAME
                ws_mas.column_dimensions['H'].width = 35  # MEET_NAME

            # FOREIGN UNMATCHED sheet (purple headers - need to add to foreign_athletes table)
            if foreign_unmatched:
                ws_foreign = wb.create_sheet(title="UNMATCHED - FOREIGN")
                ws_foreign.append(unmatched_headers)
                foreign_fill = PatternFill(start_color="9933FF", end_color="9933FF", fill_type="solid")  # Purple
                for cell in ws_foreign[1]:
                    cell.fill = foreign_fill
                    cell.font = header_font
                for missing in foreign_unmatched:
                    ws_foreign.append([
                        missing.get('sheet', ''),
                        missing.get('row', ''),
                        missing.get('full_name', ''),
                        missing.get('birthdate', ''),
                        missing.get('gender', ''),
                        missing.get('excel_nation', ''),
                        missing.get('club_name', ''),
                        missing.get('meet_name', ''),
                    ])
                # Column widths
                ws_foreign.column_dimensions['A'].width = 15  # SHEET
                ws_foreign.column_dimensions['B'].width = 6   # ROW
                ws_foreign.column_dimensions['C'].width = 35  # FULLNAME
                ws_foreign.column_dimensions['D'].width = 12  # BIRTHDATE
                ws_foreign.column_dimensions['E'].width = 8   # GENDER
                ws_foreign.column_dimensions['F'].width = 10  # EXCEL_NATION
                ws_foreign.column_dimensions['G'].width = 30  # CLUB_NAME
                ws_foreign.column_dimensions['H'].width = 35  # MEET_NAME

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Return as Excel file download with summary headers for frontend
        from fastapi.responses import StreamingResponse
        filename = f"swimrankings_preview.xlsx"

        # Print summary to backend console for visibility
        print(f"\n{'='*60}")
        print(f"[SWIMRANKINGS PREVIEW] SUMMARY")
        print(f"  Total Results: {total_rows}")
        print(f"  MATCHED (will upload): {matched_count}")
        print(f"  UNMATCHED (will NOT upload): {unmatched_count}")
        print(f"{'='*60}\n")

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "X-Preview-Total": str(total_rows),
                "X-Preview-Matched": str(matched_count),
                "X-Preview-Unmatched": str(unmatched_count),
                "Access-Control-Expose-Headers": "X-Preview-Total, X-Preview-Matched, X-Preview-Unmatched"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Preview generation failed: {str(e)}")

    finally:
        # Clean up temporary file
        try:
            os.unlink(temp_file_path)
        except:
            pass


@router.post("/admin/test-seag-upload")
async def test_seag_upload(file: UploadFile = File(...)):
    """TEST SEAG upload - Check athlete matching WITHOUT writing to database"""
    print(f"\n[TEST SEAG] Received: {file.filename}")

    # Validate file type
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")

    # Save uploaded file temporarily
    print(f"[TEST SEAG] Saving file to temporary location...")
    with tempfile.NamedTemporaryFile(delete=False, suffix=Path(file.filename).suffix) as temp_file:
        temp_file_path = temp_file.name
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            temp_file.write(chunk)

    print(f"[TEST SEAG] File saved, starting test...")

    try:
        import pandas as pd

        # Read Excel file
        try:
            df = pd.read_excel(temp_file_path, sheet_name="Sheet", skiprows=[0], header=0)
        except Exception as e:
            return {
                "success": False,
                "message": f"Failed to read Excel file: {str(e)}",
                "matched_count": 0,
                "unmatched_count": 0,
                "invalid_count": 0,
                "matched_athletes": [],
                "unmatched_athletes": [],
                "invalid_rows": []
            }

        print(f"[TEST SEAG] Loaded {len(df)} rows from Excel")

        # Get database connection (READ ONLY for testing)
        conn = get_database_connection()
        cursor = conn.cursor()

        # Import the normalize_name function for word extraction
        from ..utils.name_matcher import normalize_name

        # Test data matching
        matched_athletes = []
        unmatched_athletes = []
        invalid_rows = []

        for idx, row in df.iterrows():
            try:
                # Extract data
                gender = str(row.get('GENDER', '')).strip().upper()[0] if pd.notna(row.get('GENDER')) else None
                distance = int(row.get('DISTANCE', 0)) if pd.notna(row.get('DISTANCE')) else None
                stroke = str(row.get('STROKE', '')).strip().upper() if pd.notna(row.get('STROKE')) else None
                fullname = str(row.get('FULLNAME', '')).strip() if pd.notna(row.get('FULLNAME')) else None
                time_str = str(row.get('SWIMTIME', '')).strip() if pd.notna(row.get('SWIMTIME')) else None
                year_age = int(row.get('AGE', 0)) if pd.notna(row.get('AGE')) else None

                # Validate required fields
                if not all([gender, distance, stroke, fullname, time_str]):
                    invalid_rows.append({
                        'row': idx + 2,
                        'fullname': fullname or '(missing)',
                        'gender': gender or '(missing)',
                        'distance': distance,
                        'stroke': stroke or '(missing)',
                        'time': time_str or '(missing)',
                        'reason': 'Missing required field'
                    })
                    continue

                # Look up athlete using centralized name matcher (READ ONLY)
                # Uses word-based matching by default (most sophisticated)
                athlete_id = match_athlete_by_name(conn, fullname, gender)

                if not athlete_id:
                    csv_norm = normalize_name(fullname)
                    csv_words = set(w for w in csv_norm.split() if w.strip())
                    unmatched_athletes.append({
                        'row': idx + 2,
                        'fullname': fullname,
                        'gender': gender,
                        'distance': distance,
                        'stroke': stroke,
                        'time': time_str,
                        'age': year_age,
                        'csv_words': csv_words
                    })
                    print(f"  [NOT FOUND] Row {idx+2}: {fullname} ({gender}) - {distance} {stroke}")
                    print(f"    CSV Words: {csv_words}", flush=True)
                else:
                    matched_athletes.append({
                        'row': idx + 2,
                        'fullname': fullname,
                        'gender': gender,
                        'distance': distance,
                        'stroke': stroke,
                        'time': time_str,
                        'athlete_id': athlete_id
                    })
                    print(f"  [MATCH] Row {idx+2}: {fullname} ({gender}) - {distance} {stroke}")

            except Exception as e:
                invalid_rows.append({
                    'row': idx + 2,
                    'fullname': row.get('FULLNAME', '(unknown)'),
                    'reason': f'Error: {str(e)}'
                })

        # Print detailed info for unmatched athletes
        if unmatched_athletes:
            print(f"\n[UNMATCHED DETAIL] === {len(unmatched_athletes)} UNMATCHED ATHLETES ===", flush=True)
            for unmatched in unmatched_athletes:
                print(f"\n[UNMATCHED] Row {unmatched['row']}: {unmatched['fullname']} ({unmatched['gender']})", flush=True)
                print(f"  CSV Words: {unmatched['csv_words']}", flush=True)

                # Try to find similar athletes in database by searching for last name
                name_parts = unmatched['fullname'].split()
                if name_parts:
                    search_term = name_parts[-1] if len(name_parts) > 1 else name_parts[0]
                    cursor.execute(
                        "SELECT id, FULLNAME FROM athletes WHERE UPPER(FULLNAME) LIKE ? AND GENDER = ? LIMIT 3",
                        (f"%{search_term.upper()}%", unmatched['gender'])
                    )
                    potential_matches = cursor.fetchall()
                    if potential_matches:
                        print(f"  Potential DB Matches:", flush=True)
                        for match_id, match_name in potential_matches:
                            match_norm = normalize_name(match_name)
                            match_words = set(w for w in match_norm.split() if w.strip())
                            print(f"    ID {match_id}: {match_name}", flush=True)
                            print(f"      DB Words: {match_words}", flush=True)
                    else:
                        print(f"  No DB matches found for search term: '{search_term}'", flush=True)

        conn.close()

        print(f"\n[TEST SEAG] Results: {len(matched_athletes)} matched, {len(unmatched_athletes)} unmatched, {len(invalid_rows)} invalid")

        return {
            "success": True,
            "message": f"Test complete: {len(matched_athletes)} matched, {len(unmatched_athletes)} unmatched",
            "matched_count": len(matched_athletes),
            "unmatched_count": len(unmatched_athletes),
            "invalid_count": len(invalid_rows),
            "matched_athletes": matched_athletes,
            "unmatched_athletes": unmatched_athletes,
            "invalid_rows": invalid_rows,
            "note": "NO CHANGES MADE TO DATABASE - This is a diagnostic test only"
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Test failed: {str(e)}")

    finally:
        # Clean up temporary file
        try:
            os.unlink(temp_file_path)
        except:
            pass


@router.get("/admin/meets")
async def get_meets():
    """Get list of all meets"""
    conn = get_database_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT
                m.id,
                m.meet_name,
                m.meet_alias as alias,
                m.meet_date as date,
                m.meet_city,
                COUNT(r.id) as result_count,
                m.meet_type as category
            FROM meets m
            LEFT JOIN results r ON m.id = r.meet_id
            GROUP BY m.id, m.meet_name, m.meet_alias, m.meet_date, m.meet_city, m.meet_type
            ORDER BY m.meet_date DESC
        """)
        meets_data = cursor.fetchall()
        
        meets = []
        for row in meets_data:
            meets.append({
                "id": row[0],
                "name": row[1],
                "alias": row[2] or "",
                "date": row[3] or "",
                "city": row[4] or "",
                "result_count": row[5] or 0,
                "category": row[6] or ""
            })
        
        return {"meets": meets}
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@router.put("/admin/meets/{meet_id}/alias")
async def update_meet_alias(meet_id: str, alias_data: AliasUpdate):
    """
    Update the alias/meet_type for a specific meet.
    
    Args:
        meet_id: UUID of the meet to update
        alias_data: AliasUpdate model with 'alias' field containing the new alias/meet_type value
    
    Returns:
        JSON response with success message
    
    Raises:
        HTTPException 404: If meet not found
        HTTPException 500: If database error occurs
    """
    conn = get_database_connection()
    try:
        cursor = conn.cursor()
        
        # Verify meet exists
        cursor.execute("SELECT id, meet_name FROM meets WHERE id = ?", (meet_id,))
        meet = cursor.fetchone()
        if not meet:
            raise HTTPException(status_code=404, detail="Meet not found")

        new_alias = alias_data.alias.strip() if alias_data.alias else ''
        
        # Update meet_alias in database
        cursor.execute("UPDATE meets SET meet_alias = ? WHERE id = ?", (new_alias, meet_id))
        conn.commit()
        
        return {
            "success": True,
            "message": f"Successfully updated alias for meet '{meet[1]}' to '{new_alias}'."
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error updating alias: {str(e)}")
    finally:
        conn.close()

@router.put("/admin/meets/{meet_id}/category")
async def update_meet_category(meet_id: str, category_data: CategoryUpdate):
    """
    Update the category for a specific meet.

    Constructs category code from participant_type and scope:
    - MAST-I (Masters International)
    - MAST-D (Masters Domestic)
    - MAST-N (Masters National Team)
    - PARA-I (Para International)
    - PARA-D (Para Domestic)
    - PARA-N (Para National Team)
    - OPEN-I (Open International)
    - OPEN-D (Open Domestic)
    - OPEN-N (Open National Team)
    """
    # Validate inputs
    valid_types = ['MAST', 'PARA', 'OPEN']
    valid_scopes = ['I', 'D', 'N']

    if category_data.participant_type not in valid_types:
        raise HTTPException(status_code=400, detail=f"Invalid participant_type. Must be one of: {valid_types}")
    if category_data.scope not in valid_scopes:
        raise HTTPException(status_code=400, detail=f"Invalid scope. Must be one of: {valid_scopes}")

    # Construct category code
    category_code = f"{category_data.participant_type}-{category_data.scope}"

    conn = get_database_connection()
    try:
        cursor = conn.cursor()

        # Verify meet exists
        cursor.execute("SELECT id, meet_name FROM meets WHERE id = ?", (meet_id,))
        meet = cursor.fetchone()
        if not meet:
            raise HTTPException(status_code=404, detail="Meet not found")

        # Update meet_type in database (stores OPEN-D, PARA-I, etc.)
        cursor.execute("UPDATE meets SET meet_type = ? WHERE id = ?", (category_code, meet_id))
        conn.commit()

        return {
            "success": True,
            "message": f"Successfully updated category for meet '{meet[1]}' to '{category_code}'."
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error updating category: {str(e)}")
    finally:
        conn.close()

@router.delete("/admin/meets/{meet_id}")
async def delete_meet(meet_id: str):
    """
    Delete a meet and all associated results.
    
    This function:
    1. Deletes all results associated with the meet
    2. Deletes the meet record itself
    3. Note: Athletes and events are NOT deleted (they may be used by other meets)
    
    Args:
        meet_id: UUID of the meet to delete
    
    Returns:
        JSON response with success message
    """
    conn = get_database_connection()
    try:
        cursor = conn.cursor()
        
        # Verify meet exists
        cursor.execute("SELECT id, meet_name FROM meets WHERE id = ?", (meet_id,))
        meet = cursor.fetchone()
        if not meet:
            raise HTTPException(status_code=404, detail="Meet not found")

        meet_name = meet[1]
        
        # Delete all results for this meet
        cursor.execute("DELETE FROM results WHERE meet_id = ?", (meet_id,))
        results_deleted = cursor.rowcount
        
        # Delete the meet
        cursor.execute("DELETE FROM meets WHERE id = ?", (meet_id,))
        conn.commit()
        
        return {
            "success": True,
            "message": f"Successfully deleted meet '{meet_name}' and {results_deleted} associated results."
        }
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error deleting meet: {str(e)}")
    finally:
        conn.close()

@router.get("/admin/meets/{meet_id}/pdf")
async def get_meet_pdf(meet_id: str):
    """
    Generate HTML report for a specific meet (printable PDF-like format).
    
    This endpoint:
    1. Retrieves all results for the specified meet from the database
    2. Groups results by event (distance + stroke + gender)
    3. Generates formatted HTML with meet information and results tables
    4. Returns HTML that can be viewed in browser or printed to PDF
    
    The HTML report includes:
    - Meet details (name, type, date, location, city)
    - Results grouped by event
    - For each result: Place, Athlete Name, Time, Age, Course
    
    Args:
        meet_id: UUID of the meet to generate report for
    
    Returns:
        HTMLResponse with formatted meet results report
    
    Raises:
        HTTPException 404: If meet not found
        HTTPException 500: If database error occurs
    """
    conn = get_database_connection()
    try:
        cursor = conn.cursor()
        
        # Get meet info
        cursor.execute("SELECT id, meet_name, meet_type, meet_date, location, meet_city FROM meets WHERE id = ?", (meet_id,))
        meet = cursor.fetchone()
        if not meet:
            raise HTTPException(status_code=404, detail="Meet not found")

        meet_name = meet[1] or ""
        meet_type = meet[2] or ""
        meet_date = meet[3] or ""
        location = meet[4] or ""
        city = meet[5] or ""
        
        # Get all results for this meet
        # First, detect actual column names in the database
        cursor.execute("PRAGMA table_info(athletes)")
        athlete_columns = {row[1]: row[1] for row in cursor.fetchall()}
        athlete_gender_col = next((col for col in athlete_columns.values() if col.lower() == "gender"), "Gender")
        
        # Build query with detected column names
        # Note: Gender should come from events table (e.gender) as it represents the event's gender category
        query = f"""
            SELECT
                a.FULLNAME as athlete_name,
                COALESCE(e.gender, a."{athlete_gender_col}") as gender,
                e.event_distance,
                e.event_stroke,
                r.time_string,
                r.comp_place,
                r.day_age,
                CASE
                    WHEN r.year_age IS NOT NULL THEN r.year_age
                    WHEN a.BIRTHDATE IS NOT NULL AND r.meet_date IS NOT NULL THEN
                        CAST(substr(r.meet_date,1,4) AS INTEGER) - CAST(substr(a.BIRTHDATE,1,4) AS INTEGER)
                    ELSE NULL
                END as year_age,
                r.meet_course
            FROM results r
            LEFT JOIN athletes a ON r.athlete_id = a.id
            LEFT JOIN events e ON r.event_id = e.id
            WHERE r.meet_id = ?
              AND r.athlete_id IS NOT NULL
              AND r.event_id IS NOT NULL
            ORDER BY e.event_distance, e.event_stroke, e.gender, r.time_seconds ASC
        """
        cursor.execute(query, (meet_id,))
        results = cursor.fetchall()
        print(f"\n[PDF GENERATION] Meet ID: {meet_id}")
        print(f"[PDF GENERATION] Found {len(results)} results from database query")
        
        # Group results by event
        events_dict = {}
        for row in results:
            # row structure: (athlete_name, gender, distance, stroke, time_string, comp_place, day_age, year_age, course)
            # Handle None values for event grouping
            distance = row[2] if row[2] is not None else 0
            stroke = row[3] if row[3] is not None else ''
            gender = row[1] if row[1] is not None else ''
            
            event_key = f"{distance}{stroke}{gender}"  # distance + stroke + gender
            if event_key not in events_dict:
                events_dict[event_key] = {
                    'distance': distance,
                    'stroke': stroke,
                    'gender': gender,
                    'results': [],
                    'seen': {}  # normalized_key -> index in results
                }
            athlete_name = row[0] or ''
            time_string = row[4] or ''
            
            # Skip only if athlete_name is truly empty - otherwise include all results
            if not athlete_name or not athlete_name.strip():
                continue
            
            # Simply append all results - no duplicate filtering
            # All results from the database query should be displayed
            events_dict[event_key]['results'].append({
                'athlete_name': athlete_name,
                'time_string': time_string,
                'comp_place': row[5],
                'day_age': row[6],
                'year_age': row[7],
                'meet_course': row[8]
            })
        
        # Generate fixed-width text format HTML
        # Format: Place (4 chars) + space + Name (30 chars) + space + Time (8 chars) + newline
        html = f"""
 <!DOCTYPE html>
 <html>
 <head>
     <meta charset="UTF-8">
     <title>{meet_name} - Results</title>
     <style>
        body {{ 
            font-family: system-ui, -apple-system, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
            margin: 20px; 
            font-size: 12pt;
            line-height: 1.4;
        }}
        h1 {{ 
            color: #2c3e50; 
            margin-bottom: 20px; 
            font-size: 14pt;
        }}
        .columns {{
            column-count: 2;
            column-gap: 24px;
        }}
        .event-section {{ 
            margin-bottom: 16px; 
            page-break-inside: avoid;
            break-inside: avoid;
            -webkit-column-break-inside: avoid;
        }}
        .event-title {{ 
            font-size: 14pt; 
            font-weight: bold; 
            margin-bottom: 0;
            padding-bottom: 0;
            color: #34495e;
        }}
        .results-text {{
            font-family: 'Courier New', monospace;
            white-space: pre;
            font-size: 11pt;
            line-height: 1.0;
            margin-top: 0;
            padding-top: 0;
        }}
        @media print {{ 
            body {{ margin: 0; }}
            .event-section {{ page-break-inside: avoid; }}
        }}
     </style>
 </head>
 <body>
     <h1>{meet_name} {meet_date}</h1>
     <div class="columns">
"""
        
        # Stroke normalization uses global normalize_stroke() function
        # Database format: Free, Back, Breast, Fly, Medley
        
        # Gender mapping - handle various formats
        gender_map = {
            'M': "Men's",
            'F': "Women's",
            'MALE': "Men's",
            'FEMALE': "Women's",
            'MEN': "Men's",
            'WOMEN': "Women's"
        }
        
        # Custom sorting function for events
        def event_sort_key(event_key):
            """Sort by: 1) Gender (M before F), 2) Stroke (Free, Back, Breast, Fly, Medley), 3) Distance (short to long)"""
            event_data = events_dict[event_key]
            stroke = event_data.get('stroke')
            distance = event_data.get('distance')
            gender = event_data.get('gender')

            # Handle None values
            stroke_str = (stroke or '').upper() if stroke is not None else ''
            gender_str = (gender or '').upper() if gender is not None else ''

            # Convert distance to int for proper numeric sorting (short to long)
            try:
                distance_int = int(distance) if distance is not None else 0
            except (ValueError, TypeError):
                distance_int = 0

            # Stroke order: Free=0, Back=1, Breast=2, Fly=3, Medley=4
            # Events table uses stroke names: "Free", "Back", "Breast", "Fly", "Medley"
            stroke_lower = stroke_str.lower()

            # Determine stroke priority based on common patterns
            if stroke_lower.startswith('free') or stroke_lower in ('fr', 'fre', 'freestyle'):
                stroke_priority = 0  # Free
            elif stroke_lower.startswith('back') or stroke_lower in ('bk', 'bac', 'backstroke'):
                stroke_priority = 1  # Back
            elif stroke_lower.startswith('breast') or stroke_lower in ('br', 'bre', 'breaststroke'):
                stroke_priority = 2  # Breast
            elif stroke_lower.startswith('fly') or stroke_lower.startswith('butter') or stroke_lower in ('bu', 'fly', 'butterfly'):
                stroke_priority = 3  # Fly
            elif stroke_lower.startswith('medley') or stroke_lower in ('me', 'im', 'med', 'medley'):
                stroke_priority = 4  # Medley
            else:
                stroke_priority = 5  # Unknown strokes go last

            # Gender priority: M=0 (Men's first), F=1 (Women's second)
            gender_priority = 0 if gender_str == 'M' else 1

            # Return sort tuple: (gender, stroke, distance)
            # This sorts: Men's before Women's, then Free/Back/Breast/Fly/Medley, then short to long distance
            return (gender_priority, stroke_priority, distance_int)
        
        # Sort only the keys using the key function
        sorted_event_keys = sorted(events_dict.keys(), key=event_sort_key)
        for event_key in sorted_event_keys:
            event_data = events_dict[event_key]
            # Get course from first result (all results in same event have same course)
            course = event_data['results'][0]['meet_course'] if event_data['results'] else 'LCM'
            course_str = course if course else 'LCM'
            
            # Format: "Men's 50m Freestyle LCM"
            stroke_val = event_data.get('stroke') or ''
            # Get gender and normalize - handle None, empty strings, and whitespace
            raw_gender = event_data.get('gender')
            if raw_gender:
                gender_val = str(raw_gender).strip().upper()
            else:
                gender_val = ''
            
            distance_val = event_data.get('distance') or 0
            
            # Normalize stroke using global function - handles any input format
            try:
                stroke_full = normalize_stroke(stroke_val) if stroke_val else 'Unknown'
            except ValueError:
                stroke_full = stroke_val or 'Unknown'
            
            # Map gender - check mappings with normalized value
            if gender_val:
                # Try exact match first
                gender_full = gender_map.get(gender_val)
                # If no match, try first character
                if not gender_full and len(gender_val) > 0:
                    gender_full = gender_map.get(gender_val[0])
            else:
                gender_full = None
            
            # Default to 'Unknown' if no mapping found
            if not gender_full:
                gender_full = 'Unknown'
            
            event_title = f"{gender_full} {distance_val}m {stroke_full} {course_str}"
            
            html += f"""    <div class="event-section">
        <div class="event-title">{event_title}</div>
        <div class="results-text">"""
            for idx, result in enumerate(event_data['results'], start=1):
                # Compute meet-specific place within this event (1..N), ignoring stored place from source
                comp_place_str = str(idx).rjust(4)
                name_str = (result['athlete_name'] or '').ljust(30)[:30]  # Truncate or pad to 30 chars
                time_str = (result['time_string'] or '').rjust(8)[:8]  # Right-align to 8 chars
                age_val = result.get('year_age')
                age_str = (str(age_val).rjust(2) if age_val is not None else '  ')
                
                html += f"{comp_place_str} {name_str} {age_str}  {time_str}\n"
            
            html += """</div>
    </div>
"""
        
        html += """
    </div>
</body>
</html>
"""
        
        return HTMLResponse(content=html)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

async def process_uploaded_file(file_path: str, filename: str, meet_name: str, meet_code: str, existing_meet_id: str = None) -> ConversionResult:
    """Process uploaded Excel file and add to database"""
    
    print(f"\n{'='*60}", flush=True)
    print(f"[UPLOAD] Starting processing: {filename}", flush=True)
    print(f"[UPLOAD] Meet: {meet_name or 'N/A'}, Code: {meet_code or 'N/A'}", flush=True)
    print(f"{'='*60}", flush=True)
    
    # Determine if we're creating a new meet or adding to existing
    conn = get_database_connection()
    try:
        cursor = conn.cursor()
        
        if existing_meet_id:
            # Add to existing meet
            cursor.execute("SELECT * FROM meets WHERE id = ?", (existing_meet_id,))
            existing_meet = cursor.fetchone()
            if not existing_meet:
                raise HTTPException(status_code=404, detail="Existing meet not found")
            
            meet_info = {
                'id': existing_meet_id,
                'name': existing_meet[1],  # name
                'meet_type': existing_meet[2],  # meet_type
                'meet_date': existing_meet[3],  # meet_date
                'location': existing_meet[4],  # location
                'city': existing_meet[5] if len(existing_meet) > 5 else None,  # city
            }
        else:
            # Create new meet (will be updated with extracted info during processing)
            # Note: Do NOT set meet_type (alias) automatically - aliases must be set manually
            meet_info = {
                'id': str(uuid.uuid4()),
                'name': meet_name or 'Uploaded Meet',
                'meet_type': None,  # Don't auto-assign aliases - must be set manually
                'meet_date': datetime.now().strftime('%Y-%m-%dT%H:%M:%SZ'),  # ISO 8601 with UTC timezone
                'location': 'Uploaded Meet',
            }
    finally:
        conn.close()
    
    # Process the file using the conversion script with full validation
    # Note: We now continue processing even with validation errors and report issues in summary
    print(f"[PARSE] Reading Excel file...", flush=True)
    file_path_obj = Path(file_path)
    validation_issues = None
    try:
        athletes, results, events, collector = process_meet_file_simple(file_path_obj, meet_info)
        validation_issues = collector
        print(f"[PARSE] OK Parsed: {len(athletes)} athletes, {len(results)} results, {len(events)} events", flush=True)
        if validation_issues and hasattr(validation_issues, 'has_errors') and validation_issues.has_errors():
            error_count = (
                len(validation_issues.missing_athletes) +
                len(getattr(validation_issues, 'name_format_mismatches', [])) +
                len(validation_issues.birthdate_mismatches) +
                len(validation_issues.nation_mismatches) +
                len(validation_issues.club_misses) +
                len(validation_issues.event_misses)
            )
            print(f"[PARSE] WARNING Found {error_count} validation issues (will be reported in summary)")

        # BLOCK UPLOAD if there are unmatched athletes
        if validation_issues and hasattr(validation_issues, 'missing_athletes') and validation_issues.missing_athletes:
            unmatched_count = len(validation_issues.missing_athletes)
            # Get sample of unmatched names (first 5)
            sample_names = [ma.get('full_name', 'Unknown') for ma in validation_issues.missing_athletes[:5]]
            sample_str = ', '.join(sample_names)
            if unmatched_count > 5:
                sample_str += f" ... and {unmatched_count - 5} more"
            print(f"[UPLOAD] BLOCKED - {unmatched_count} unmatched athletes found")
            return ConversionResult(
                success=False,
                message=f"Upload blocked: {unmatched_count} unmatched athlete(s) found. Run PREVIEW first to identify and add missing athletes before uploading. Unmatched: {sample_str}"
            )
    except Exception as e:
        # If there's a different error, still raise it
        print(f"[PARSE] ERROR parsing file: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Conversion failed: {str(e)}")
    
    # After processing, check if meet with extracted name already exists (deduplication)
    if not existing_meet_id and meet_info.get('name'):
        conn = get_database_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT id, meet_name, meet_date, meet_city, meet_type FROM meets WHERE meet_name = ?", (meet_info['name'],))
            existing_meet = cursor.fetchone()
            
            if existing_meet:
                # Use existing meet ID instead of creating duplicate
                meet_info['id'] = existing_meet[0]
                # Update meet_info with existing meet data
                if existing_meet[4]:  # meet_type
                    meet_info['meet_type'] = existing_meet[4]
                if existing_meet[2]:  # meet_date
                    meet_info['meet_date'] = existing_meet[2]
                if existing_meet[3]:  # city
                    meet_info['city'] = existing_meet[3]
        finally:
            conn.close()
    
    
    if not results:
        print(f"[UPLOAD] ERROR No results found in file")
        return ConversionResult(
            success=False,
            message="No valid swimming results found in the Excel file. Please check the file format."
        )
    
    # Insert into database: split results by meet_name to create separate meets
    print(f"[DB] Starting database insertion...")
    conn = get_database_connection()
    try:
        def norm_text(s: str) -> str:
            if s is None:
                return ''
            s = str(s).strip()
            return ' '.join(s.split())
        results_by_meet = {}
        mattioli_debug = []  # Track Mattioli results specifically
        for r in results:
            raw_name = r.get('meet_name') or meet_info.get('name') or 'Uploaded Meet'
            name_key = norm_text(raw_name)
            results_by_meet.setdefault(name_key, []).append(r)
            # Debug: track any result that might be Mattioli
            if 'mattioli' in raw_name.lower() or 'victorian' in raw_name.lower():
                mattioli_debug.append({
                    'raw_name': raw_name,
                    'normalized': name_key,
                    'sheet': r.get('sheet_name', '?'),
                    'row': r.get('excel_row', '?'),
                    'athlete': r.get('full_name', '?')
                })


        total_meets_created = 0
        per_meet_summaries = []
        print(f"[DB] Processing {len(results_by_meet)} meet group(s)...")
        for idx, (name, group) in enumerate(results_by_meet.items(), 1):
            print(f"[DB] [{idx}/{len(results_by_meet)}] Processing meet: '{name}' ({len(group)} results)")
            # Determine earliest date and city from group
            # Earliest date
            raw_dates = [r.get('meet_date') for r in group if r.get('meet_date')]
            parsed = []
            for d in raw_dates:
                for fmt in ('%Y-%m-%d','%d/%m/%Y','%d.%m.%Y','%Y/%m/%d','%d-%m-%Y'):
                    try:
                        parsed.append(datetime.strptime(str(d), fmt).strftime('%Y-%m-%d'))
                        break
                    except Exception:
                        pass
            earliest_date = min(parsed) if parsed else meet_info.get('meet_date')
            cities = [norm_text(r.get('meet_city')) for r in group if r.get('meet_city')]
            city = cities[0] if cities else norm_text(meet_info.get('city'))

            # Reuse existing meet if same name, city, and year exists
            # Note: Use year only (not full date) since meets span multiple days
            # Matching by full date causes men's and women's files for the same meet to not match
            # But we need the year to distinguish same meet name in different years
            cursor = conn.cursor()
            year = None
            if earliest_date:
                try:
                    # Extract year from earliest_date (format: YYYY-MM-DD or YYYY)
                    year = earliest_date[:4] if len(earliest_date) >= 4 else None
                except Exception:
                    pass
            
            # Match by name + city + year (extract year from meet_date in database)
            existing = None
            if city and year:
                # Extract year from meet_date column for comparison
                cursor.execute("""
                    SELECT id, meet_date, meet_city, meet_type FROM meets
                    WHERE meet_name = ? AND meet_city = ?
                    AND substr(meet_date, 1, 4) = ?
                """, (name, city, year))
                existing = cursor.fetchone()
            # Fallback: name + city only (if year missing)
            if not existing and city:
                cursor.execute("SELECT id, meet_date, meet_city, meet_type FROM meets WHERE meet_name = ? AND meet_city = ?",
                             (name, city))
                existing = cursor.fetchone()
            # Last resort: name only (if city is missing)
            if not existing:
                cursor.execute("SELECT id, meet_date, meet_city, meet_type FROM meets WHERE meet_name = ?", (name,))
                existing = cursor.fetchone()
            
            if existing:
                # Use existing meet - preserve existing alias (meet_type), don't overwrite it
                existing_alias = existing[3]  # meet_type
                child_meet_info = {
                    'id': existing[0],
                    'name': name,
                    'meet_type': existing_alias,  # Preserve existing alias, don't overwrite
                    'meet_date': existing[1] or earliest_date,  # Use existing date if available
                    'location': meet_info.get('location'),
                    'city': existing[2] or city,  # Use existing city if available
                }
            else:
                # Create new meet - don't set alias automatically
                new_meet_id = str(uuid.uuid4())
                child_meet_info = {
                    'id': new_meet_id,
                    'name': name,
                    'meet_type': None,  # Don't auto-assign aliases
                    'meet_date': earliest_date,
                    'location': meet_info.get('location'),
                    'city': city,
                }

            # Update meet_id on group results to the child meet id
            meet_id_assigned = child_meet_info['id']
            for r in group:
                r['meet_id'] = meet_id_assigned

            # Insert and collect summary (pass collector to track skipped rows)
            print(f"[DB]   Inserting {len(group)} results into database...")
            summary = insert_data_simple(conn, athletes, group, events, child_meet_info, collector=validation_issues)
            inserted = summary.get('inserted_results', 0)
            skipped = summary.get('skipped_results', 0)
            print(f"[DB]   OK Inserted: {inserted} new, {skipped} duplicates skipped")
            per_meet_summaries.append((name, child_meet_info['meet_date'], child_meet_info.get('city'), summary))
            total_meets_created += 0 if existing else 1

        # Build summary message with clear statistics
        lines = [f"Upload complete: {filename}\n"]
        total_inserted = 0
        total_skipped = 0
        for name, d, city, s in per_meet_summaries:
            ins = s.get('inserted_results', 0); sk = s.get('skipped_results', 0)
            total_inserted += ins
            total_skipped += sk
            city_str = f", {city}" if city else ""
            if ins > 0 or sk > 0:
                lines.append(f"  - {name} ({d}{city_str}): {ins} results added, {sk} duplicate results skipped (already in database)")
        
        # Add totals summary - emphasize duplicates and skipped rows
        if total_skipped > 0:
            lines.append(f"\nSUMMARY: {total_inserted} new results inserted, {total_skipped} duplicate results skipped (already in database)")
        else:
            lines.append(f"\nSUMMARY: {total_inserted} results inserted, no duplicates found")
        
        # Report validation issues as informational (not errors)
        # Group by meet for better readability
        if validation_issues and hasattr(validation_issues, 'has_errors') and validation_issues.has_errors():
            lines.append("\n" + "="*60)
            lines.append("VALIDATION ISSUES BY MEET")
            lines.append("="*60)
            
            # Create a mapping of meet names to their validation issues
            meets_with_issues = {}
            
            # Helper to get meet name from an issue
            def get_meet_name(issue, default="Unknown Meet"):
                # Try to get from meet_name field first
                meet = issue.get("meet_name", "")
                if meet and meet != "(unknown)":
                    return meet
                # Fallback to sheet name
                sheet = issue.get("sheet", "")
                if sheet:
                    # Try to match sheet name to a meet in per_meet_summaries
                    for name, _, _, _ in per_meet_summaries:
                        if name.lower() in sheet.lower() or sheet.lower() in name.lower():
                            return name
                return default
            
            # Group missing athletes by meet
            if validation_issues.missing_athletes:
                for athlete in validation_issues.missing_athletes:
                    meet_name = get_meet_name(athlete)
                    if meet_name not in meets_with_issues:
                        meets_with_issues[meet_name] = {"missing_athletes": [], "birthdate_mismatches": [], "nation_mismatches": [], "name_format_mismatches": [], "club_misses": [], "event_misses": []}
                    meets_with_issues[meet_name]["missing_athletes"].append(athlete)
            
            # Group name format mismatches by meet
            if hasattr(validation_issues, 'name_format_mismatches') and validation_issues.name_format_mismatches:
                for mismatch in validation_issues.name_format_mismatches:
                    meet_name = get_meet_name(mismatch)
                    if meet_name not in meets_with_issues:
                        meets_with_issues[meet_name] = {"missing_athletes": [], "birthdate_mismatches": [], "nation_mismatches": [], "name_format_mismatches": [], "club_misses": [], "event_misses": []}
                    meets_with_issues[meet_name]["name_format_mismatches"].append(mismatch)
            
            # Group birthdate mismatches by meet
            if validation_issues.birthdate_mismatches:
                for mismatch in validation_issues.birthdate_mismatches:
                    meet_name = get_meet_name(mismatch)
                    if meet_name not in meets_with_issues:
                        meets_with_issues[meet_name] = {"missing_athletes": [], "birthdate_mismatches": [], "nation_mismatches": [], "name_format_mismatches": [], "club_misses": [], "event_misses": []}
                    meets_with_issues[meet_name]["birthdate_mismatches"].append(mismatch)
            
            # Group nation mismatches by meet
            if validation_issues.nation_mismatches:
                for mismatch in validation_issues.nation_mismatches:
                    meet_name = get_meet_name(mismatch, "Other Issues")
                    if meet_name not in meets_with_issues:
                        meets_with_issues[meet_name] = {"missing_athletes": [], "birthdate_mismatches": [], "nation_mismatches": [], "name_format_mismatches": [], "club_misses": [], "event_misses": []}
                    meets_with_issues[meet_name]["nation_mismatches"].append(mismatch)
            
            # Group club misses by meet
            if validation_issues.club_misses:
                for miss in validation_issues.club_misses:
                    meet_name = get_meet_name(miss)
                    if meet_name not in meets_with_issues:
                        meets_with_issues[meet_name] = {"missing_athletes": [], "birthdate_mismatches": [], "nation_mismatches": [], "name_format_mismatches": [], "club_misses": [], "event_misses": []}
                    meets_with_issues[meet_name]["club_misses"].append(miss)
            
            # Group event misses by meet
            if validation_issues.event_misses:
                for miss in validation_issues.event_misses:
                    meet_name = get_meet_name(miss)
                    if meet_name not in meets_with_issues:
                        meets_with_issues[meet_name] = {"missing_athletes": [], "birthdate_mismatches": [], "nation_mismatches": [], "name_format_mismatches": [], "club_misses": [], "event_misses": []}
                    meets_with_issues[meet_name]["event_misses"].append(miss)
            
            # Display issues grouped by meet
            for meet_name in sorted(meets_with_issues.keys()):
                issues = meets_with_issues[meet_name]
                has_any = any(issues.values())
                if not has_any:
                    continue
                
                lines.append(f"\n{meet_name}:")
                lines.append("-" * 60)
                
                # Missing athletes - NEED TO BE ADDED
                if issues["missing_athletes"]:
                    lines.append(f"  [WARNING] MISSING ATHLETES - NEED TO BE ADDED TO DATABASE ({len(issues['missing_athletes'])} rows):")
                    for athlete in issues["missing_athletes"][:20]:  # First 20
                        sheet = athlete.get("sheet", "Unknown")
                        row = athlete.get("row", "?")
                        full_name = athlete.get("full_name", "Unknown")
                        birthdate = athlete.get("birthdate", "(blank)")
                        gender = athlete.get("gender", "(blank)")
                        lines.append(f"    - {sheet} row {row}: {full_name} (Birthdate: {birthdate}, Gender: {gender}) - ADD TO DATABASE")
                    if len(issues["missing_athletes"]) > 20:
                        lines.append(f"    ... and {len(issues['missing_athletes']) - 20} more missing athletes that need to be added")
                    lines.append("")
                    lines.append("  [ACTION REQUIRED] These athletes are not in the database.")
                    lines.append("     Add them to the athlete table before uploading results.")
                
                # Birthdate mismatches
                if issues["birthdate_mismatches"]:
                    lines.append(f"  Birthdate Mismatches ({len(issues['birthdate_mismatches'])} rows):")
                    for mismatch in issues["birthdate_mismatches"][:20]:  # First 20
                        sheet = mismatch.get("sheet", "Unknown")
                        row = mismatch.get("row", "?")
                        full_name = mismatch.get("full_name", "Unknown")
                        excel_bday = mismatch.get("excel_birthdate", "(blank)")
                        db_bday = mismatch.get("database_birthdate", "(blank)")
                        athlete_id = mismatch.get("athlete_id", "(unknown)")
                        lines.append(f"    - {sheet} row {row}: {full_name} (athlete_id: {athlete_id}) - Excel: {excel_bday} != Database: {db_bday}")
                    if len(issues["birthdate_mismatches"]) > 20:
                        lines.append(f"    ... and {len(issues['birthdate_mismatches']) - 20} more birthdate mismatches")
                
                # Nation mismatches
                if issues["nation_mismatches"]:
                    lines.append(f"  Nation Mismatches ({len(issues['nation_mismatches'])} rows):")
                    for mismatch in issues["nation_mismatches"][:20]:  # First 20
                        sheet = mismatch.get("sheet", "Unknown")
                        row = mismatch.get("row", "?")
                        full_name = mismatch.get("full_name", "Unknown")
                        excel_nation = mismatch.get("excel_nation", "(blank)")
                        db_nation = mismatch.get("database_nation", "(blank)")
                        athlete_id = mismatch.get("athlete_id", "(unknown)")
                        lines.append(f"    - {sheet} row {row}: {full_name} (athlete_id: {athlete_id}) - Excel: {excel_nation} != Database: {db_nation}")
                    if len(issues["nation_mismatches"]) > 20:
                        lines.append(f"    ... and {len(issues['nation_mismatches']) - 20} more nation mismatches")
                
                # Name format mismatches (exact FULLNAME doesn't match but normalized does)
                if issues.get("name_format_mismatches"):
                    lines.append(f"  Name Format Mismatches ({len(issues['name_format_mismatches'])} rows - case/format differences):")
                    for mismatch in issues["name_format_mismatches"][:20]:  # First 20
                        sheet = mismatch.get("sheet", "Unknown")
                        row = mismatch.get("row", "?")
                        upload_name = mismatch.get("upload_fullname", "Unknown")
                        db_name = mismatch.get("database_fullname", "Unknown")
                        lines.append(f"    - {sheet} row {row}: Upload='{upload_name}' != Database='{db_name}'")
                    if len(issues["name_format_mismatches"]) > 20:
                        lines.append(f"    ... and {len(issues['name_format_mismatches']) - 20} more name format mismatches")
                
                # Club misses
                if issues["club_misses"]:
                    lines.append(f"  Unknown Clubs/Teams ({len(issues['club_misses'])} rows - results created with null club fields):")
                    for miss in issues["club_misses"][:20]:  # First 20
                        sheet = miss.get("sheet", "Unknown")
                        row = miss.get("row", "?")
                        full_name = miss.get("full_name", "Unknown")
                        club_name = miss.get("club_name", "(blank)")
                        athlete_id = miss.get("athlete_id", "(unknown)")
                        lines.append(f"    - {sheet} row {row}: {full_name} (athlete_id: {athlete_id}) - Club not found: {club_name}")
                    if len(issues["club_misses"]) > 20:
                        lines.append(f"    ... and {len(issues['club_misses']) - 20} more unknown clubs")
                
                # Event misses
                if issues["event_misses"]:
                    lines.append(f"  Unknown Events ({len(issues['event_misses'])} rows):")
                    for miss in issues["event_misses"][:20]:  # First 20
                        sheet = miss.get("sheet", "Unknown")
                        row = miss.get("row", "?")
                        description = miss.get("description", "Unknown event")
                        lines.append(f"    - {sheet} row {row}: {description}")
                    if len(issues["event_misses"]) > 20:
                        lines.append(f"    ... and {len(issues['event_misses']) - 20} more unknown events")
                
                # Skipped rows (duplicates and other reasons) - only show if there are any
                if hasattr(validation_issues, 'skipped_rows') and validation_issues.skipped_rows:
                    # Group skipped rows by reason for this meet
                    skipped_by_reason = {}
                    for skipped in validation_issues.skipped_rows:
                        reason = skipped.get("reason", "Unknown reason")
                        if reason not in skipped_by_reason:
                            skipped_by_reason[reason] = []
                        # Only include if it's for this meet (check sheet name matches meet name pattern)
                        skipped_sheet = skipped.get("sheet", "")
                        if not meet_name or meet_name == "Unknown Meet" or meet_name.lower() in skipped_sheet.lower() or skipped_sheet.lower() in meet_name.lower():
                            skipped_by_reason[reason].append(skipped)
                    
                    if skipped_by_reason:
                        total_skipped_for_meet = sum(len(rows) for rows in skipped_by_reason.values())
                        lines.append(f"\n  Skipped Rows ({total_skipped_for_meet} rows - not uploaded):")
                        for reason, rows in skipped_by_reason.items():
                            lines.append(f"    - {reason}: {len(rows)} row(s)")
                            # Show first 5 examples of each reason
                            for skipped in rows[:5]:
                                sheet = skipped.get("sheet", "Unknown")
                                row = skipped.get("row", "?")
                                full_name = skipped.get("full_name", "Unknown")
                                lines.append(f"      - {sheet} row {row}: {full_name}")
                            if len(rows) > 5:
                                lines.append(f"      ... and {len(rows) - 5} more rows with this reason")
            
            lines.append("\n" + "="*60)
        
        message = "\n".join(lines)
        
        # Final summary
        total_inserted = sum(s.get('inserted_results', 0) for _, _, _, s in per_meet_summaries)
        total_skipped = sum(s.get('skipped_results', 0) for _, _, _, s in per_meet_summaries)
        print(f"\n{'='*60}")
        print(f"[UPLOAD] COMPLETE!")
        print(f"[UPLOAD]   Total inserted: {total_inserted} results")
        print(f"[UPLOAD]   Total skipped (duplicates): {total_skipped} results")
        print(f"[UPLOAD]   Meets processed: {len(results_by_meet)}")
        print(f"{'='*60}\n")
        
        # Extract unmatched clubs and name format mismatches from validation collector
        unmatched_clubs_list = []
        club_misses_list = []
        name_format_mismatches_list = []
        if validation_issues:
            if hasattr(validation_issues, 'club_misses'):
                # Get unique unmatched club names (for backward compatibility)
                seen_clubs = set()
                for club_miss in validation_issues.club_misses:
                    club_name = club_miss.get('club_name', '').strip()
                    if club_name and club_name not in seen_clubs and club_name != '(blank)':
                        seen_clubs.add(club_name)
                        unmatched_clubs_list.append({
                            'club_name': club_name,
                            'state_code': None  # Could extract from club name if needed
                        })
                    # Also include full club miss data with context
                    if club_name and club_name != '(blank)':
                        club_misses_list.append(club_miss)
            if hasattr(validation_issues, 'name_format_mismatches'):
                # Get name format mismatches with athlete_id
                for mismatch in validation_issues.name_format_mismatches:
                    if mismatch.get('athlete_id'):
                        name_format_mismatches_list.append(mismatch)
        
        # Extract missing athletes from validation collector
        missing_athletes_list = []
        if validation_issues and hasattr(validation_issues, 'missing_athletes'):
            missing_athletes_list = validation_issues.missing_athletes
            
            # Automatically save missing athletes to JSON file for easy access
            if missing_athletes_list:
                import json
                output_file = project_root / "missing_athletes_latest.json"
                try:
                    with open(output_file, 'w', encoding='utf-8') as f:
                        json.dump({
                            "timestamp": datetime.now().isoformat(),
                            "filename": filename,
                            "missing_athletes": missing_athletes_list
                        }, f, indent=2, ensure_ascii=False)
                    print(f"[SAVE] Saved {len(missing_athletes_list)} missing athletes to: {output_file}", flush=True)
                except Exception as e:
                    print(f"[WARN] Failed to save missing athletes to file: {e}", flush=True)
        
        return ConversionResult(
            success=True,
            message=message,
            athletes=len(athletes),
            results=len(results),
            events=len(events),
            meets=max(1, len(results_by_meet)),
            unmatched_clubs=unmatched_clubs_list,
            club_misses=club_misses_list,
            name_format_mismatches=name_format_mismatches_list,
            missing_athletes=missing_athletes_list,
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Conversion failed: {str(e)}")
    finally:
        conn.close()

@router.post("/admin/convert-clubs", response_model=ClubConversionResult)
async def convert_clubs_excel(file: UploadFile = File(...)):
    """Convert uploaded Clubs_By_State.xlsx file to database entries"""
    
    print(f"[club upload] Received request for file: {file.filename}")
    
    # Validate file type
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")
    
    # Save uploaded file temporarily by streaming
    temp_file_path = None
    try:
        suffix = Path(file.filename).suffix
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp_file:
            temp_file_path = temp_file.name
            print(f"[club upload] Received file '{file.filename}', starting save to '{temp_file_path}'")
            # Stream file in chunks
            chunk_size = 1024 * 1024  # 1MB
            total_saved = 0
            while True:
                chunk = await file.read(chunk_size)
                if not chunk:
                    break
                temp_file.write(chunk)
                total_saved += len(chunk)
            print(f"[club upload] Saved {total_saved} bytes for '{file.filename}'")
        
        # Process the file
        print(f"[club upload] Processing '{file.filename}'...")
        try:
            # Quick validation before processing - check file structure
            import pandas as pd
            quick_check = pd.ExcelFile(temp_file_path, engine='xlrd' if suffix == '.xls' else None)
            sheet_names_check = quick_check.sheet_names[:5]  # Check first 5 sheets
            
            # Check if this looks like a meet results file
            meet_indicators = ['50M', '100M', '200M', '400M', '800M', '1500M', 'FR', 'BK', 'BR', 'BU', 'ME', 'IM']
            has_meet_indicators = any(any(ind in str(s).upper() for ind in meet_indicators) for s in sheet_names_check)
            state_code_sheets = [s for s in sheet_names_check if len(str(s).strip().upper()) >= 2 and len(str(s).strip().upper()) <= 3]
            
            if has_meet_indicators and len(state_code_sheets) == 0:
                raise HTTPException(status_code=400, detail="This appears to be a meet results file (has event sheets like '50m Fr', '100m Fr'), not a Clubs_By_State file. Please upload this to the 'Upload & Convert Meet' tab instead. Clubs_By_State files should have sheets named with state codes (e.g., 'KL', 'SGR', 'JHR').")
            
            data = process_clubs_file(temp_file_path)
            print(f"[club upload] Processed: {len(data.get('states', []))} states, {len(data.get('clubs', []))} clubs")
        except HTTPException:
            # Re-raise HTTP exceptions as-is
            raise
        except ValueError as ve:
            # This is a file structure validation error
            raise HTTPException(status_code=400, detail=str(ve))
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")
        
        # Insert into database
        conn = get_database_connection()
        try:
            result = insert_club_data(conn, data['states'], data['clubs'])
            print(f"[club upload] Inserted: {result['inserted_states']} states, {result['inserted_clubs']} clubs")
            print(f"[club upload] Skipped: {result['skipped_states']} states, {result['skipped_clubs']} clubs")
            
            message = (f"Successfully converted {file.filename}.\n"
                      f"{result['inserted_states']} states added, {result['skipped_states']} states skipped\n"
                      f"{result['inserted_clubs']} clubs added, {result['skipped_clubs']} clubs skipped")
            
            return ClubConversionResult(
                success=True,
                message=message,
                states=len(data['states']),
                clubs=len(data['clubs']),
                inserted_states=result['inserted_states'],
                skipped_states=result['skipped_states'],
                inserted_clubs=result['inserted_clubs'],
                skipped_clubs=result['skipped_clubs']
            )
        finally:
            conn.close()
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Conversion failed: {str(e)}")
    
    finally:
        # Clean up temporary file
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
                print(f"[club upload] Cleaned up temporary file: {temp_file_path}")
            except Exception as e:
                print(f"[club upload] Error cleaning up temporary file {temp_file_path}: {e}")


@router.get("/admin/athletes/search")
async def search_athletes(q: str = ""):
    """Search athletes by name and aliases (case-insensitive).
    Uses core search function from global name matcher for consistency."""
    query = (q or "").strip()
    if not query:
        return {"athletes": []}

    # Use the core search function from name_matcher - single source of truth
    try:
        from src.web.utils.name_matcher import search_athletes_by_name
    except ImportError:
        # Fallback if import fails
        return {"athletes": [], "error": "Name matcher not available"}

    conn = get_database_connection()
    try:
        athletes = search_athletes_by_name(conn, query, limit=50, include_aliases=True)
        return {"athletes": athletes}
    finally:
        conn.close()


@router.get("/admin/events/export-excel")
async def export_events_excel():
    """Export ALL columns from events table as Excel file."""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from datetime import datetime

    try:
        conn = get_database_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM events ORDER BY event_course, gender, event_stroke, event_distance")
        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        conn.close()

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Events"

        # Add headers (actual column names from database)
        ws.append(columns)

        # Style headers (red background, white text)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Add data rows
        for row in rows:
            ws.append(list(row))

        # Auto-adjust column widths
        for col_idx, col_name in enumerate(columns, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            max_width = min(max(len(str(col_name)) + 2, 10), 50)
            ws.column_dimensions[col_letter].width = max_width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"events_full_export_{timestamp}.xlsx"

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export events: {str(e)}")


@router.get("/admin/events/filter")
async def filter_events(course: str = None, gender: str = None):
    """
    Filter events by course and/or gender.

    Parameters:
    - course: Filter by course (LCM, SCM, YD) - required first
    - gender: Filter by gender (M, F, X) - optional, narrows results further

    Returns: List of event objects with all properties needed for display/editing
    """
    conn = get_database_connection()
    cursor = conn.cursor()

    # Build query - filter by course first, then gender if provided
    query = "SELECT id, event_distance, event_stroke, gender, event_course, event_type FROM events WHERE 1=1"
    params = []

    if course:
        query += " AND event_course = ?"
        params.append(course)

    if gender:
        query += " AND gender = ?"
        params.append(gender)

    query += " ORDER BY id"

    cursor.execute(query, params)
    events = cursor.fetchall()
    cursor.close()
    conn.close()

    # Convert to list of objects (keep database format, no reformatting)
    event_list = []
    for row in events:
        event_list.append({
            "id": row[0],
            "distance": row[1],
            "stroke": row[2],
            "gender": row[3],
            "course": row[4],
            "event_type": row[5]
        })

    return {"events": event_list}


@router.options("/admin/events/{event_id}")
async def event_options(event_id: str):
    """CORS preflight handler for event updates."""
    return {}


@router.patch("/admin/events/{event_id}")
async def update_event(event_id: str, distance: int = None, stroke: str = None, gender: str = None, course: str = None):
    """
    Update an event's fields.

    Supports updating:
    - distance: Event distance in meters (50, 100, 200, 400, 800, 1500)
    - stroke: Event stroke (Fr, Bk, Br, Bu, Me) - accepts any case format
    - gender: Event gender (M, F, X)
    - course: Event course (LCM, SCM, YD)

    Event ID is regenerated if distance, stroke, or gender changes.
    Duplicate ID check prevents conflicts.
    """
    from datetime import datetime

    if not any([distance, stroke, gender, course]):
        raise HTTPException(status_code=400, detail="At least one field to update is required")

    try:
        conn = get_database_connection()
        cursor = conn.cursor()

        # Get current event
        cursor.execute("SELECT id, event_distance, event_stroke, gender, event_course FROM events WHERE id = ?", (event_id,))
        current_event = cursor.fetchone()

        if not current_event:
            conn.close()
            raise HTTPException(status_code=404, detail=f"Event not found: {event_id}")

        # Use current values if not updating
        new_distance = distance if distance is not None else current_event[1]
        new_stroke = stroke if stroke is not None else current_event[2]
        new_gender = gender if gender is not None else current_event[3]
        new_course = course if course is not None else current_event[4]

        # If stroke provided, normalize it
        if stroke is not None:
            try:
                new_stroke = normalize_stroke(stroke)
            except ValueError as e:
                conn.close()
                raise HTTPException(status_code=400, detail=str(e))

        # Check if regenerating ID (distance, stroke, or gender changed)
        id_changed = (distance is not None and distance != current_event[1]) or \
                     (stroke is not None and stroke != current_event[2]) or \
                     (gender is not None and gender != current_event[3])

        new_event_id = event_id
        if id_changed:
            # Generate new ID: {COURSE}_{STROKE}_{DISTANCE}_{GENDER}
            new_event_id = f"{new_course}_{new_stroke}_{new_distance}_{new_gender}"

            # Check for duplicates
            cursor.execute("SELECT id FROM events WHERE id = ? AND id != ?", (new_event_id, event_id))
            if cursor.fetchone():
                conn.close()
                raise HTTPException(
                    status_code=409,
                    detail=f"Event ID already exists: {new_event_id}. Cannot update to create duplicate."
                )

        # Update the event
        if id_changed:
            # If ID changed: delete old, insert new (to maintain referential integrity)
            # First, get the event_type before deleting
            cursor.execute("SELECT event_type FROM events WHERE id = ?", (event_id,))
            event_type_result = cursor.fetchone()
            event_type = event_type_result[0] if event_type_result else "Individual"

            # Update any results that reference this event
            cursor.execute("UPDATE results SET event_id = ? WHERE event_id = ?", (new_event_id, event_id))

            # Delete old event
            cursor.execute("DELETE FROM events WHERE id = ?", (event_id,))

            # Insert new event with preserved event_type
            cursor.execute("""
                INSERT INTO events (id, event_distance, event_stroke, gender, event_course, event_type, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (new_event_id, new_distance, new_stroke, new_gender, new_course, event_type, datetime.now().isoformat() + 'Z'))
        else:
            # If only non-ID fields changed, just update
            update_fields = []
            update_values = []

            if distance is not None:
                update_fields.append("event_distance = ?")
                update_values.append(new_distance)

            if stroke is not None:
                update_fields.append("event_stroke = ?")
                update_values.append(new_stroke)

            if gender is not None:
                update_fields.append("gender = ?")
                update_values.append(new_gender)

            if course is not None:
                update_fields.append("event_course = ?")
                update_values.append(new_course)

            if update_fields:
                update_values.append(event_id)
                query = f"UPDATE events SET {', '.join(update_fields)} WHERE id = ?"
                cursor.execute(query, update_values)

        conn.commit()
        conn.close()

        message = f"Event updated successfully"
        if id_changed:
            message += f". Event ID changed from {event_id} to {new_event_id}"

        return {
            "success": True,
            "message": message,
            "event_id": new_event_id,
            "distance": new_distance,
            "stroke": new_stroke,
            "gender": new_gender,
            "course": new_course,
            "id_changed": id_changed
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to update event: {str(e)}")


@router.get("/admin/results/export-excel")
async def export_results_excel():
    """Export all results from the database as an exact replica of the results table."""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from datetime import datetime
    import sys

    try:
        conn = get_database_connection()
        cursor = conn.cursor()

        # Fetch all columns and data from the results table (CORRECT QUERY)
        cursor.execute("SELECT * FROM results")
        
        results = cursor.fetchall()
        
        # Get column names (headers) from the cursor description
        if not results:
            # If the table is empty, use the known headers
            headers = [
                'ID', 'Meet ID', 'Athlete ID', 'Event ID', 'Time (Seconds)', 'Time (String)', 
                'Time (Seconds Numeric)', 'comp_place', 'Aqua Points', 'Rudolph Points', 
                'Course', 'Result Meet Date', 'Day Age', 'Year Age', 'Created At', 
                'Team Name', 'Team Code', 'Team State Code', 'Team Nation', 
                'Is Relay'
            ]
        else:
            # Use the actual column names returned by the database (which use underscores, e.g., 'time_seconds')
            headers = [description[0].replace('_', ' ').title() for description in cursor.description]

        conn.close()

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results Data"

        # Add headers
        ws.append(headers)

        # Style header row (using the same red/white style as the athlete export)
        header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Add data rows
        for row in results:
            ws.append(row)

        # Adjust widths for readability (optional, based on column count)
        for i, column in enumerate(ws.columns, 1):
            length = max(len(str(cell.value or '')) for cell in column)
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = length + 2

        # Write to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"results_export_{timestamp}.xlsx"

        # Return as StreamingResponse
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        import traceback
        print(f"DEBUG: Exception occurred in results export: {str(e)}", flush=True)
        traceback.print_exc(file=sys.stdout)
        sys.stdout.flush()
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")
    
@router.get("/admin/athletes/export-excel")
async def export_athletes_excel():
    """Export ALL columns from athletes table as Excel file"""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from datetime import datetime

    try:
        conn = get_database_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM athletes ORDER BY FULLNAME ASC")
        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        conn.close()

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Athletes"

        # Add headers (actual column names from database)
        ws.append(columns)

        # Style header row
        header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Add data rows
        for row in rows:
            ws.append(list(row))

        # Auto-adjust column widths based on header
        for col_idx, col_name in enumerate(columns, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            max_width = min(max(len(str(col_name)) + 2, 10), 50)
            ws.column_dimensions[col_letter].width = max_width

        # Write to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"athletes_full_export_{timestamp}.xlsx"

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@router.get("/admin/foreign-athletes/export-excel")
async def export_foreign_athletes_excel():
    """Export all foreign athletes from database as Excel file"""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from datetime import datetime

    try:
        conn = get_database_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM foreign_athletes ORDER BY fullname ASC")
        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Foreign Athletes"

        ws.append(columns)
        header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for row in rows:
            ws.append(list(row))

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"foreign_athletes_export_{timestamp}.xlsx"

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@router.get("/admin/coaches/export-excel")
async def export_coaches_excel():
    """Export all coaches from database as Excel file"""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from datetime import datetime

    try:
        conn = get_database_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM coaches ORDER BY coach_name ASC")
        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Coaches"

        ws.append(columns)
        header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for row in rows:
            ws.append(list(row))

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"coaches_export_{timestamp}.xlsx"

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@router.get("/admin/clubs/export-excel")
async def export_clubs_excel():
    """Export all clubs from database as Excel file"""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from datetime import datetime

    try:
        conn = get_database_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM clubs ORDER BY club_name ASC")
        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clubs"

        ws.append(columns)
        header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for row in rows:
            ws.append(list(row))

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"clubs_export_{timestamp}.xlsx"

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@router.get("/admin/athletes/{athlete_id}")
async def get_athlete_detail(athlete_id: str):
    """Get full athlete details including all fields."""
    conn = get_database_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM athletes WHERE id = ?", (athlete_id,))
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Athlete not found")

        # Get column names
        columns = [description[0] for description in cursor.description]

        # Build athlete dict with all fields
        athlete = {}
        for i, col in enumerate(columns):
            athlete[col] = row[i] if row[i] is not None else ""

        # Also add mapped names for frontend compatibility
        athlete["name"] = athlete.get("FULLNAME", "")
        athlete["gender"] = athlete.get("Gender", "")
        athlete["birth_date"] = athlete.get("BIRTHDATE", "")

        # If athlete has club_code but no state_code, infer state from club
        if athlete.get("club_code") and not athlete.get("state_code"):
            cursor.execute("SELECT state_code FROM clubs WHERE club_code = ?", (athlete["club_code"],))
            club_row = cursor.fetchone()
            if club_row and club_row[0]:
                athlete["state_code"] = club_row[0]

        return {"athlete": athlete}
    finally:
        conn.close()


@router.options("/admin/athletes/{athlete_id}")
async def athlete_options(athlete_id: str):
    """Handle CORS preflight for athlete update endpoint"""
    return {}

@router.patch("/admin/athletes/{athlete_id}")
async def update_athlete(athlete_id: str, payload: AthleteUpdateRequest):
    conn = get_database_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM athletes WHERE id = ?", (athlete_id,))
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="Athlete not found")

        # Build dynamic UPDATE query based on provided fields
        update_fields = []
        update_values = []

        # Map API field names to database column names
        field_name_map = {
            'ClubCode': 'club_code',
            'ClubName': 'club_name',
            'NATION': 'nation',
            'nation': 'nation',  # lowercase alias
            'club_code': 'club_code',  # lowercase alias
            'club_name': 'club_name',  # lowercase alias
            'state_code': 'state_code',  # state code
            'postal_code': 'postal_code',  # postal code
            'address_state': 'address_state',  # address state (separate from club state)
        }

        payload_dict = payload.model_dump(exclude_none=False)

        for field_name, field_value in payload_dict.items():
            if field_value is not None:
                # Strip whitespace and handle case sensitivity for specific fields
                if isinstance(field_value, str):
                    cleaned_value = field_value.strip()
                    # Uppercase specific fields
                    if field_name in ['NATION', 'nation', 'Gender', 'state_code']:
                        cleaned_value = cleaned_value.upper() if cleaned_value else None
                else:
                    cleaned_value = field_value

                # Map API field name to database column name
                db_column_name = field_name_map.get(field_name, field_name)
                update_fields.append(f"{db_column_name} = ?")
                update_values.append(cleaned_value if cleaned_value else None)

        if not update_fields:
            raise HTTPException(status_code=400, detail="No fields to update")

        update_values.append(athlete_id)

        sql_query = f"""
            UPDATE athletes
            SET {', '.join(update_fields)}
            WHERE id = ?
        """

        cursor.execute(sql_query, update_values)
        conn.commit()
        return {"success": True, "message": "Athlete updated"}
    except HTTPException:
        raise
    except Exception as exc:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to update athlete: {exc}")
    finally:
        conn.close()

@router.get("/admin/clubs/unmatched")
async def get_unmatched_clubs():
    """Get list of unmatched clubs from the last upload (stored in validation collector)"""
    # For now, we'll query results table for clubs that weren't matched
    # In the future, we could store unmatched clubs in a separate table
    conn = get_database_connection()
    cursor = conn.cursor()
    
    try:
        # Get unique club names from results where club_name is NULL (unmatched)
        # But we need to track what the original club name was
        # For now, let's get clubs that appear in results but not in clubs table
        cursor.execute("""
            SELECT DISTINCT r.club_name as unmatched_name, r.state_code
            FROM results r
            LEFT JOIN clubs c ON UPPER(TRIM(r.club_name)) = UPPER(TRIM(c.club_name))
            WHERE r.club_name IS NOT NULL
            AND r.club_name != ''
            AND c.club_name IS NULL
            ORDER BY r.club_name
        """)
        
        unmatched = []
        for row in cursor.fetchall():
            unmatched.append({
                "club_name": row[0],
                "state_code": row[1] if row[1] else None
            })
        
        return {"unmatched_clubs": unmatched, "count": len(unmatched)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get unmatched clubs: {str(e)}")
    finally:
        conn.close()

@router.post("/admin/clubs")
async def create_club(club: ClubCreateRequest):
    """Create a new club in the database"""
    conn = get_database_connection()
    cursor = conn.cursor()
    
    try:
        # Check if club already exists (case-insensitive)
        cursor.execute(
            "SELECT club_name FROM clubs WHERE UPPER(TRIM(club_name)) = ?",
            (club.club_name.upper().strip(),)
        )
        if cursor.fetchone():
            raise HTTPException(status_code=400, detail=f"Club '{club.club_name}' already exists")
        
        # Check if alias column exists
        cursor.execute("PRAGMA table_info(clubs)")
        columns = {row[1].lower() for row in cursor.fetchall()}
        has_alias = 'club_alias' in columns
        
        # Insert new club
        if has_alias:
            cursor.execute("""
                INSERT INTO clubs (club_name, club_code, state_code, club_nation, club_alias)
                VALUES (?, ?, ?, ?, ?)
            """, (
                club.club_name.strip(),
                club.club_code.strip() if club.club_code else None,
                club.state_code.strip().upper() if club.state_code else None,
                club.nation.strip().upper() if club.nation else "MAS",
                club.alias.strip() if club.alias else None
            ))
        else:
            cursor.execute("""
                INSERT INTO clubs (club_name, club_code, state_code, club_nation)
                VALUES (?, ?, ?, ?)
            """, (
                club.club_name.strip(),
                club.club_code.strip() if club.club_code else None,
                club.state_code.strip().upper() if club.state_code else None,
                club.nation.strip().upper() if club.nation else "MAS"
            ))
        
        conn.commit()
        return {"success": True, "message": f"Club '{club.club_name}' created successfully"}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create club: {str(e)}")
    finally:
        conn.close()

@router.get("/admin/clubs")
async def list_clubs(state_code: Optional[str] = None):
    """List all clubs in the database, optionally filtered by state_code"""
    conn = get_database_connection()
    cursor = conn.cursor()
    
    try:
        # Check if alias column exists
        cursor.execute("PRAGMA table_info(clubs)")
        columns = {row[1].lower() for row in cursor.fetchall()}
        has_alias = 'club_alias' in columns
        
        # Build query with optional state filter
        if state_code:
            if has_alias:
                cursor.execute(
                    "SELECT club_name, club_code, state_code, club_nation, club_alias FROM clubs WHERE state_code = ? ORDER BY club_name",
                    (state_code.upper(),)
                )
            else:
                cursor.execute(
                    "SELECT club_name, club_code, state_code, club_nation FROM clubs WHERE state_code = ? ORDER BY club_name",
                    (state_code.upper(),)
                )
        else:
            if has_alias:
                cursor.execute("SELECT club_name, club_code, state_code, club_nation, club_alias FROM clubs ORDER BY club_name")
            else:
                cursor.execute("SELECT club_name, club_code, state_code, club_nation FROM clubs ORDER BY club_name")
        
        clubs = []
        for row in cursor.fetchall():
            club = {
                "club_name": row[0],
                "club_code": row[1],
                "state_code": row[2],
                "nation": row[3]
            }
            if has_alias and len(row) > 4:
                club["alias"] = row[4]
            clubs.append(club)
        
        return {"clubs": clubs, "count": len(clubs)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to list clubs: {str(e)}")
    finally:
        conn.close()

@router.get("/admin/clubs/states")
async def list_states():
    """Get list of unique state codes from clubs"""
    conn = get_database_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("SELECT DISTINCT state_code FROM clubs WHERE state_code IS NOT NULL AND state_code != '' ORDER BY state_code")
        states = [row[0] for row in cursor.fetchall()]
        return {"states": states}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to list states: {str(e)}")
    finally:
        conn.close()

@router.put("/admin/clubs/{club_name}")
async def update_club(club_name: str, club: ClubCreateRequest):
    """Update an existing club in the database"""
    conn = get_database_connection()
    cursor = conn.cursor()
    
    try:
        # Check if club exists
        cursor.execute(
            "SELECT club_name FROM clubs WHERE club_name = ?",
            (club_name,)
        )
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail=f"Club '{club_name}' not found")

        # Check if club_code is already in use by another club
        if club.club_code:
            cursor.execute(
                "SELECT club_name FROM clubs WHERE club_code = ? AND club_name != ?",
                (club.club_code.strip().upper(), club_name)
            )
            existing = cursor.fetchone()
            if existing:
                raise HTTPException(
                    status_code=409,
                    detail=f"Club code '{club.club_code.strip().upper()}' is already in use by '{existing[0]}'"
                )

        # Check if alias column exists
        cursor.execute("PRAGMA table_info(clubs)")
        columns = {row[1].lower() for row in cursor.fetchall()}
        has_alias = 'club_alias' in columns
        
        # Update club - use actual column names: club_nation, club_alias
        if has_alias:
            cursor.execute("""
                UPDATE clubs
                SET club_name = ?, club_code = ?, state_code = ?, club_nation = ?, club_alias = ?
                WHERE club_name = ?
            """, (
                club.club_name.strip(),
                club.club_code.strip() if club.club_code else None,
                club.state_code.strip().upper() if club.state_code else None,
                club.nation.strip().upper() if club.nation else "MAS",
                club.alias.strip() if club.alias else None,
                club_name  # WHERE clause uses original name
            ))
        else:
            cursor.execute("""
                UPDATE clubs
                SET club_name = ?, club_code = ?, state_code = ?, club_nation = ?
                WHERE club_name = ?
            """, (
                club.club_name.strip(),
                club.club_code.strip() if club.club_code else None,
                club.state_code.strip().upper() if club.state_code else None,
                club.nation.strip().upper() if club.nation else "MAS",
                club_name  # WHERE clause uses original name
            ))
        
        conn.commit()
        return {"success": True, "message": f"Club '{club.club_name}' updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to update club: {str(e)}")
    finally:
        conn.close()

@router.delete("/admin/clubs/{club_name}")
async def delete_club(club_name: str):
    """Delete a club by name"""
    conn = get_database_connection()
    cursor = conn.cursor()

    try:
        # Check if club exists
        cursor.execute("SELECT club_name FROM clubs WHERE club_name = ?", (club_name,))
        existing = cursor.fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail=f"Club '{club_name}' not found")

        # Check if club has associated results
        cursor.execute("SELECT COUNT(*) FROM results WHERE club_name = ?", (club_name,))
        result_count = cursor.fetchone()[0]
        if result_count > 0:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot delete club '{club_name}' - it has {result_count} associated results. Remove results first."
            )

        # Check if club has associated athletes
        cursor.execute("SELECT COUNT(*) FROM athletes WHERE club_name = ?", (club_name,))
        athlete_count = cursor.fetchone()[0]
        if athlete_count > 0:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot delete club '{club_name}' - it has {athlete_count} associated athletes. Reassign athletes first."
            )

        # Delete the club
        cursor.execute("DELETE FROM clubs WHERE club_name = ?", (club_name,))
        conn.commit()

        return {"success": True, "message": f"Club '{club_name}' deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete club: {str(e)}")
    finally:
        conn.close()

@router.get("/admin/coaches/search")
async def search_coaches(q: str = ""):
    """Search coaches by name (word-based matching)"""
    if len(q) < 2:
        return {"coaches": []}

    conn = get_database_connection()
    cursor = conn.cursor()

    try:
        # Split query into words for matching
        words = q.upper().strip().split()

        # Build query to find coaches matching any words
        query = """
            SELECT id, coach_name, birthdate, gender, nation, club_name, coach_role, state_coach, state_code, msn_program
            FROM coaches
            WHERE 1=1
        """
        params = []

        for word in words:
            query += " AND UPPER(coach_name) LIKE ?"
            params.append(f"%{word}%")

        query += " ORDER BY coach_name LIMIT 50"

        cursor.execute(query, params)
        coaches = []
        for row in cursor.fetchall():
            coaches.append({
                "id": row[0],
                "coach_name": row[1],
                "birthdate": row[2],
                "gender": row[3],
                "nation": row[4],
                "club_name": row[5],
                "coach_role": row[6],
                "state_coach": row[7],
                "state_code": row[8],
                "msn_program": row[9],
            })

        return {"coaches": coaches}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to search coaches: {str(e)}")
    finally:
        conn.close()


@router.get("/admin/coaches/{coach_id}")
async def get_coach(coach_id: int):
    """Get a single coach by ID"""
    conn = get_database_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT * FROM coaches WHERE id = ?", (coach_id,))
        row = cursor.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Coach not found")

        # Get column names
        columns = [description[0] for description in cursor.description]
        coach = dict(zip(columns, row))

        return {"coach": coach}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get coach: {str(e)}")
    finally:
        conn.close()


@router.get("/admin/coaches")
async def list_coaches(club_name: Optional[str] = None, name: Optional[str] = None):
    """List coaches, optionally filtered by club_name or name"""
    conn = get_database_connection()
    cursor = conn.cursor()
    
    try:
        query = """
            SELECT id, club_name, name, role, email, whatsapp,
                   passport_photo, passport_number, ic, shoe_size, tshirt_size, tracksuit_size,
                   course_level_1_sport_specific, course_level_2, course_level_3,
                   course_level_1_isn, course_level_2_isn, course_level_3_isn,
                   seminar_oct_2024, other_courses, state_coach, logbook_file,
                   created_at, updated_at
            FROM coaches
            WHERE 1=1
        """
        params = []
        
        if club_name:
            query += " AND club_name = ?"
            params.append(club_name)
        
        if name:
            query += " AND name LIKE ?"
            params.append(f"%{name}%")
        
        query += " ORDER BY club_name, name"
        
        cursor.execute(query, params)
        coaches = []
        for row in cursor.fetchall():
            coaches.append({
                "id": row[0],
                "club_name": row[1],
                "name": row[2],
                "role": row[3],
                "email": row[4],
                "whatsapp": row[5],
                "passport_photo": row[6],
                "passport_number": row[7],
                "ic": row[8],
                "shoe_size": row[9],
                "tshirt_size": row[10],
                "tracksuit_size": row[11],
                "course_level_1_sport_specific": bool(row[12]),
                "course_level_2": bool(row[13]),
                "course_level_3": bool(row[14]),
                "course_level_1_isn": bool(row[15]),
                "course_level_2_isn": bool(row[16]),
                "course_level_3_isn": bool(row[17]),
                "seminar_oct_2024": bool(row[18]),
                "other_courses": row[19],
                "state_coach": bool(row[20]),
                "logbook_file": row[21],
                "created_at": row[22],
                "updated_at": row[23]
            })
        
        return {"coaches": coaches, "count": len(coaches)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to list coaches: {str(e)}")
    finally:
        conn.close()

@router.post("/admin/coaches")
async def create_coach(coach: CoachCreateRequest):
    """Create a new coach"""
    conn = get_database_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            INSERT INTO coaches (
                club_name, coach_name, coach_role, coach_email, coach_whatsapp,
                coach_passport_photo, coach_passport_number, coach_ic, shoe_size, tshirt_size, tracksuit_size,
                course_level_1_sport_specific, course_level_2, course_level_3,
                course_level_1_isn, course_level_2_isn, course_level_3_isn,
                seminar_oct_2024, other_courses, state_coach, logbook_file
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            coach.club_name.strip(),
            coach.name.strip(),
            coach.role.strip(),
            coach.email.strip() if coach.email else None,
            coach.whatsapp.strip() if coach.whatsapp else None,
            coach.passport_photo.strip() if coach.passport_photo else None,
            coach.passport_number.strip() if coach.passport_number else None,
            coach.ic.strip() if coach.ic else None,
            coach.shoe_size.strip() if coach.shoe_size else None,
            coach.tshirt_size.strip() if coach.tshirt_size else None,
            coach.tracksuit_size.strip() if coach.tracksuit_size else None,
            1 if coach.course_level_1_sport_specific else 0,
            1 if coach.course_level_2 else 0,
            1 if coach.course_level_3 else 0,
            1 if coach.course_level_1_isn else 0,
            1 if coach.course_level_2_isn else 0,
            1 if coach.course_level_3_isn else 0,
            1 if coach.seminar_oct_2024 else 0,
            coach.other_courses.strip() if coach.other_courses else None,
            1 if coach.state_coach else 0,
            coach.logbook_file.strip() if coach.logbook_file else None
        ))
        
        conn.commit()
        return {"success": True, "message": f"Coach '{coach.name}' created successfully", "id": cursor.lastrowid}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create coach: {str(e)}")
    finally:
        conn.close()

@router.put("/admin/coaches/{coach_id}")
async def update_coach(coach_id: int, data: dict):
    """Update an existing coach with flexible field mapping"""
    conn = get_database_connection()
    cursor = conn.cursor()

    try:
        # Check if coach exists
        cursor.execute("SELECT id FROM coaches WHERE id = ?", (coach_id,))
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail=f"Coach with id {coach_id} not found")

        # Build dynamic update query based on provided fields
        update_fields = []
        params = []

        # Map of frontend field names to database column names
        field_map = {
            'coach_name': 'coach_name',
            'birthdate': 'birthdate',
            'gender': 'gender',
            'nation': 'nation',
            'club_name': 'club_name',
            'coach_role': 'coach_role',
            'state_coach': 'state_coach',
            'state_code': 'state_code',
            'msn_program': 'msn_program',
            'coach_passport_number': 'coach_passport_number',
            'coach_email': 'coach_email',
            'coach_phone': 'coach_phone',
        }

        for frontend_field, db_column in field_map.items():
            if frontend_field in data:
                update_fields.append(f"{db_column} = ?")
                value = data[frontend_field]
                # Handle string stripping
                if isinstance(value, str):
                    value = value.strip() if value else None
                params.append(value)

        if not update_fields:
            return {"success": True, "message": "No fields to update"}

        params.append(coach_id)
        query = f"UPDATE coaches SET {', '.join(update_fields)} WHERE id = ?"

        cursor.execute(query, params)
        conn.commit()

        return {"success": True, "message": "Coach updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to update coach: {str(e)}")
    finally:
        conn.close()

@router.delete("/admin/coaches/{coach_id}")
async def delete_coach(coach_id: int):
    """Delete a coach"""
    conn = get_database_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("SELECT coach_name FROM coaches WHERE id = ?", (coach_id,))
        coach = cursor.fetchone()
        if not coach:
            raise HTTPException(status_code=404, detail=f"Coach with id {coach_id} not found")
        
        cursor.execute("DELETE FROM coaches WHERE id = ?", (coach_id,))
        conn.commit()
        return {"success": True, "message": f"Coach '{coach[0]}' deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete coach: {str(e)}")
    finally:
        conn.close()

@router.get("/admin/clubs/search")
async def search_clubs(query: str = Query(..., min_length=2)):
    """Search for clubs by name (case-insensitive partial match)"""
    conn = get_database_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT club_name, club_code, state_code, nation, alias
            FROM clubs
            WHERE LOWER(club_name) LIKE LOWER(?) OR LOWER(club_code) LIKE LOWER(?)
            ORDER BY club_name
            LIMIT 20
        """, (f"%{query}%", f"%{query}%"))
        
        results = cursor.fetchall()
        clubs = []
        for row in results:
            clubs.append({
                "club_name": row[0],
                "club_code": row[1],
                "state_code": row[2],
                "nation": row[3],
                "alias": row[4]
            })
        
        return {"clubs": clubs, "count": len(clubs)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to search clubs: {str(e)}")
    finally:
        conn.close()

class ClubResolutionRequest(BaseModel):
    club_miss_id: Optional[str] = None  # Identifier for the club miss (e.g., index or hash)
    action: str  # "add_alias", "swap_names", or "create_new"
    existing_club_name: Optional[str] = None  # Required for add_alias and swap_names
    new_club_name: Optional[str] = None  # Required for create_new
    club_code: Optional[str] = None
    state_code: Optional[str] = None
    nation: str = "MAS"

@router.post("/admin/clubs/resolve-miss")
async def resolve_club_miss(resolution: ClubResolutionRequest):
    """Resolve a club miss by adding alias, swapping names, or creating new club"""
    conn = get_database_connection()
    cursor = conn.cursor()
    
    try:
        # Check if alias column exists
        cursor.execute("PRAGMA table_info(clubs)")
        columns = [row[1].lower() for row in cursor.fetchall()]
        has_alias = "alias" in columns
        
        if not has_alias:
            cursor.execute("ALTER TABLE clubs ADD COLUMN alias TEXT")
        
        if resolution.action == "add_alias":
            # Add the read-in name as an alias to existing club
            if not resolution.existing_club_name:
                raise HTTPException(status_code=400, detail="existing_club_name required for add_alias action")
            
            # Get current alias
            cursor.execute("SELECT club_alias FROM clubs WHERE club_name = ?", (resolution.existing_club_name,))
            row = cursor.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail=f"Club '{resolution.existing_club_name}' not found")

            current_alias = row[0] or ""
            # The "read-in name" should be passed as new_club_name (the name from Excel)
            new_alias = resolution.new_club_name or ""

            if new_alias:
                if current_alias:
                    updated_alias = f"{current_alias}, {new_alias}"
                else:
                    updated_alias = new_alias

                cursor.execute("UPDATE clubs SET club_alias = ? WHERE club_name = ?", (updated_alias, resolution.existing_club_name))
                conn.commit()
                return {"success": True, "message": f"Added '{new_alias}' as alias to '{resolution.existing_club_name}'"}
        
        elif resolution.action == "swap_names":
            # Make read-in name the club_name and existing club_name the alias
            if not resolution.existing_club_name or not resolution.new_club_name:
                raise HTTPException(status_code=400, detail="Both existing_club_name and new_club_name required for swap_names action")
            
            # Check if new name already exists
            cursor.execute("SELECT club_name FROM clubs WHERE club_name = ?", (resolution.new_club_name,))
            if cursor.fetchone():
                raise HTTPException(status_code=400, detail=f"Club '{resolution.new_club_name}' already exists")
            
            # Get current alias
            cursor.execute("SELECT club_alias FROM clubs WHERE club_name = ?", (resolution.existing_club_name,))
            row = cursor.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail=f"Club '{resolution.existing_club_name}' not found")

            current_alias = row[0] or ""
            # Add old club_name to alias
            if current_alias:
                updated_alias = f"{current_alias}, {resolution.existing_club_name}"
            else:
                updated_alias = resolution.existing_club_name

            # Update club_name and club_alias
            cursor.execute("""
                UPDATE clubs
                SET club_name = ?, club_alias = ?
                WHERE club_name = ?
            """, (resolution.new_club_name, updated_alias, resolution.existing_club_name))
            conn.commit()
            return {"success": True, "message": f"Swapped names: '{resolution.existing_club_name}' -> '{resolution.new_club_name}' (old name added as alias)"}
        
        elif resolution.action == "create_new":
            # Create a new club
            if not resolution.new_club_name:
                raise HTTPException(status_code=400, detail="new_club_name required for create_new action")
            
            # Check if club already exists
            cursor.execute("SELECT club_name FROM clubs WHERE LOWER(club_name) = LOWER(?)", (resolution.new_club_name,))
            if cursor.fetchone():
                raise HTTPException(status_code=400, detail=f"Club '{resolution.new_club_name}' already exists")
            
            cursor.execute("""
                INSERT INTO clubs (club_name, club_code, state_code, club_nation, club_alias)
                VALUES (?, ?, ?, ?, ?)
            """, (
                resolution.new_club_name,
                resolution.club_code,
                resolution.state_code,
                resolution.nation,
                None  # No alias for new club
            ))
            conn.commit()
            return {"success": True, "message": f"Created new club '{resolution.new_club_name}'"}
        
        else:
            raise HTTPException(status_code=400, detail=f"Invalid action: {resolution.action}. Must be 'add_alias', 'swap_names', or 'create_new'")
    
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to resolve club miss: {str(e)}")
    finally:
        conn.close()

@router.post("/admin/athletes/{athlete_id}/add-alias")
async def add_athlete_alias(athlete_id: str, alias: str = Query(...)):
    """Add an alias to an athlete's alias field"""
    conn = get_database_connection()
    cursor = conn.cursor()
    
    try:
        # Check if athlete exists
        cursor.execute("SELECT id, FULLNAME FROM athletes WHERE id = ?", (athlete_id,))
        athlete = cursor.fetchone()
        if not athlete:
            raise HTTPException(status_code=404, detail=f"Athlete with id {athlete_id} not found")
        
        # Check which alias columns exist
        cursor.execute("PRAGMA table_info(athletes)")
        columns_info = cursor.fetchall()
        columns = {row[1].lower() for row in columns_info}
        has_alias_1 = "athlete_alias_1" in columns
        has_alias_2 = "athlete_alias_2" in columns
        
        if not has_alias_1 and not has_alias_2:
            raise HTTPException(status_code=400, detail="Alias columns not found in athletes table")
        
        # Check current alias values
        current_alias_1 = None
        current_alias_2 = None
        if has_alias_1 or has_alias_2:
            select_cols = []
            if has_alias_1:
                select_cols.append("athlete_alias_1")
            if has_alias_2:
                select_cols.append("athlete_alias_2")
            if select_cols:
                cursor.execute(f"SELECT {', '.join(select_cols)} FROM athletes WHERE id = ?", (athlete_id,))
                row = cursor.fetchone()
                if row:
                    idx = 0
                    if has_alias_1:
                        current_alias_1 = row[idx]
                        idx += 1
                    if has_alias_2 and len(row) > idx:
                        current_alias_2 = row[idx]
        
        # Normalize alias for comparison
        def normalize_name(name):
            if not name:
                return ""
            return str(name).strip().upper()
        
        # Check if alias already exists
        alias_normalized = normalize_name(alias)
        if (current_alias_1 and normalize_name(current_alias_1) == alias_normalized) or \
           (current_alias_2 and normalize_name(current_alias_2) == alias_normalized):
            return {"success": True, "message": f"Alias '{alias}' already exists for this athlete"}
        
        # Add alias to first available field
        if has_alias_1 and not current_alias_1:
            cursor.execute("UPDATE athletes SET athlete_alias_1 = ? WHERE id = ?", (alias.strip(), athlete_id))
            conn.commit()
            return {"success": True, "message": f"Alias '{alias}' added to athlete_alias_1"}
        elif has_alias_2 and not current_alias_2:
            cursor.execute("UPDATE athletes SET athlete_alias_2 = ? WHERE id = ?", (alias.strip(), athlete_id))
            conn.commit()
            return {"success": True, "message": f"Alias '{alias}' added to athlete_alias_2"}
        else:
            # Both aliases are full - overwrite alias_2
            cursor.execute("UPDATE athletes SET athlete_alias_2 = ? WHERE id = ?", (alias.strip(), athlete_id))
            conn.commit()
            return {"success": True, "message": f"Alias '{alias}' added to athlete_alias_2 (overwrote existing value)"}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to add alias: {str(e)}")
    finally:
        conn.close()

@router.get("/admin/athletes/{athlete_id}/results")
async def get_athlete_results(athlete_id: str, start_date: Optional[str] = None, end_date: Optional[str] = None, best_only: bool = False):
    """Get results for a specific athlete with club and state information"""
    conn = get_database_connection()
    cursor = conn.cursor()
    
    try:
        # Build query with date filtering and best times option
        base_query = """
            SELECT 
                r.id,
                r.time_string,
                r.comp_place,
                r.aqua_points,
                r.year_age,
                r.day_age,
                e.event_distance,
                e.event_stroke,
                e.gender as event_gender,
                m.id as meet_id,
                m.meet_name as meet_name,
                m.meet_type as meet_code,
                m.meet_date,
                r.club_name as club_name,
                r.state_code as state_code,
                a.FULLNAME as athlete_name,
                a.BIRTHDATE as birthdate
            FROM results r
            LEFT JOIN athletes a ON r.athlete_id = a.id
            LEFT JOIN events e ON r.event_id = e.id
            LEFT JOIN meets m ON r.meet_id = m.id
            WHERE r.athlete_id = ?
        """
        params = [athlete_id]
        
        # Add date filtering
        if start_date:
            base_query += " AND (m.meet_date IS NULL OR m.meet_date >= ?)"
            params.append(start_date)
        
        if end_date:
            base_query += " AND (m.meet_date IS NULL OR m.meet_date <= ?)"
            params.append(end_date)
        
        # For best times, we need to group by event and get the best time
        if best_only:
            base_query += """
                ORDER BY e.event_distance, e.event_stroke, e.gender,
                         COALESCE(r.time_seconds, 999999)
            """
            cursor.execute(base_query, params)
            all_results = cursor.fetchall()
            
            # Group by event and keep only best time
            best_results = {}
            for row in all_results:
                event_key = f"{row[6]}_{row[7]}_{row[8]}"  # distance_stroke_gender
                if event_key not in best_results:
                    best_results[event_key] = row
                else:
                    # Compare times (lower is better)
                    current_time = row[1] or "99:99.99"
                    best_time = best_results[event_key][1] or "99:99.99"
                    # Simple string comparison for times (works for most formats)
                    if current_time < best_time:
                        best_results[event_key] = row
            
            results = list(best_results.values())
        else:
            base_query += " ORDER BY m.meet_date DESC, e.event_distance, e.event_stroke"
            cursor.execute(base_query, params)
            results = cursor.fetchall()
        
        # Convert to list of dicts
        data = []
        for row in results:
            data.append({
                "result_id": row[0],
                "time": row[1],
                "comp_place": row[2],
                "aqua_points": row[3],
                "year_age": row[4],
                "day_age": row[5],
                "distance": row[6],
                "stroke": row[7],
                "event_gender": row[8],
                "meet_id": row[9],
                "meet_name": row[10],
                "meet_code": row[11],
                "meet_date": row[12],
                "club_name": row[13],
                "state_code": row[14],
                "athlete_name": row[15],
                "birthdate": row[16]
            })
        
        return {"results": data, "count": len(data)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get athlete results: {str(e)}")
    finally:
        conn.close()

@router.get("/admin/clubs/{club_name}/roster")
async def get_club_roster(club_name: str, meet_id: Optional[str] = None):
    """Get athlete roster for a club, optionally filtered by meet"""
    conn = get_database_connection()
    cursor = conn.cursor()
    
    try:
        # Build query to get unique athletes from results for this club
        base_query = """
            SELECT DISTINCT
                a.id as athlete_id,
                a.FULLNAME as fullname,
                a.Gender as gender,
                r.year_age,
                r.day_age,
                m.meet_date,
                a.BIRTHDATE as birthdate
            FROM results r
            LEFT JOIN athletes a ON r.athlete_id = a.id
            LEFT JOIN meets m ON r.meet_id = m.id
            WHERE r.club_name = ?
        """
        params = [club_name]
        
        if meet_id:
            base_query += " AND r.meet_id = ?"
            params.append(meet_id)
        
        base_query += " ORDER BY a.FULLNAME"
        
        cursor.execute(base_query, params)
        results = cursor.fetchall()
        
        # Convert to list of dicts and compute age
        roster = []
        for row in results:
            # Compute age at meet date
            age = None
            if row[6] and row[7]:  # meet_date and birthdate
                try:
                    from datetime import datetime
                    meet_date = datetime.strptime(row[6], '%Y-%m-%d') if isinstance(row[6], str) else row[6]
                    birthdate = datetime.strptime(row[7], '%Y.%m.%d') if isinstance(row[7], str) else row[7]
                    age = (meet_date.year - birthdate.year) - ((meet_date.month, meet_date.day) < (birthdate.month, birthdate.day))
                except:
                    age = row[3]  # fallback to year_age
            elif row[3]:  # year_age
                age = row[3]
            
            roster.append({
                "athlete_id": row[0],
                "fullname": row[1] or "Unknown",
                "gender": row[2] or "U",
                "age": age,
                "year_age": row[3],
                "day_age": row[4]
            })
        
        return {"roster": roster, "count": len(roster), "club_name": club_name}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get club roster: {str(e)}")
    finally:
        conn.close()

@router.post("/admin/analyze-athlete-info")
async def analyze_athlete_info(file: UploadFile = File(...)):
    """Analyze uploaded athlete info workbook structure - returns headers, sample data, and sheet info"""
    
    print(f"[athlete info analysis] Received request for file: {file.filename}")
    
    # Validate file type
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")
    
    import pandas as pd
    
    # Basic validation: check if this looks like a meet results file
    # (We'll do a quick check after reading the file)
    
    # Save uploaded file temporarily
    temp_file_path = None
    try:
        suffix = Path(file.filename).suffix
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp_file:
            temp_file_path = temp_file.name
            print(f"[athlete info analysis] Received file '{file.filename}', saving to '{temp_file_path}'")
            # Stream file in chunks
            chunk_size = 1024 * 1024  # 1MB
            while True:
                chunk = await file.read(chunk_size)
                if not chunk:
                    break
                temp_file.write(chunk)
        
        # Analyze the file
        print(f"[athlete info analysis] Analyzing '{file.filename}'...")
        excel_file = pd.ExcelFile(temp_file_path, engine='xlrd' if suffix == '.xls' else None)
        sheet_names = excel_file.sheet_names
        
        print(f"[athlete info analysis] Found {len(sheet_names)} sheet(s): {', '.join(sheet_names[:5])}")
        
        # Basic validation: check if this looks like a meet results file instead
        meet_indicators = ['50M', '100M', '200M', '400M', '800M', '1500M', 'FR', 'BK', 'BR', 'BU', 'ME', 'IM']
        has_meet_indicators = any(any(ind in str(s).upper() for ind in meet_indicators) for s in sheet_names[:3])
        has_state_codes = any(len(str(s).strip().upper()) >= 2 and len(str(s).strip().upper()) <= 3 for s in sheet_names[:3])
        
        if has_meet_indicators and not has_state_codes:
            raise HTTPException(status_code=400, detail="This appears to be a meet results file (has event sheets like '50m Fr', '100m Fr'). Please upload this to the 'Upload & Convert Meet' tab instead. Athlete Info files should contain athlete personal information.")
        
        sheets_info = []
        for sheet_name in sheet_names:
            try:
                # Read first few rows to get headers and sample data
                df = pd.read_excel(temp_file_path, sheet_name=sheet_name, nrows=5, engine='xlrd' if suffix == '.xls' else None)
                
                # Get headers (first row)
                headers = [str(col).strip() if pd.notna(col) else f'Column_{i}' for i, col in enumerate(df.columns)]
                
                # Get sample rows (first 3 data rows)
                sample_rows = []
                for idx in range(min(3, len(df))):
                    row = df.iloc[idx]
                    sample_rows.append([str(val).strip() if pd.notna(val) else '' for val in row.values])
                
                sheets_info.append({
                    'name': sheet_name,
                    'row_count': len(pd.read_excel(temp_file_path, sheet_name=sheet_name, engine='xlrd' if suffix == '.xls' else None)),
                    'column_count': len(df.columns),
                    'headers': headers,
                    'sample_rows': sample_rows
                })
            except Exception as e:
                print(f"Error analyzing sheet '{sheet_name}': {e}")
                sheets_info.append({
                    'name': sheet_name,
                    'error': str(e)
                })
        
        return {
            'filename': file.filename,
            'sheet_count': len(sheet_names),
            'sheets': sheets_info
        }
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Analysis failed: {str(e)}")
    
    finally:
        # Clean up temporary file
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
            except Exception as e:
                print(f"[athlete info analysis] Error cleaning up temporary file: {e}")


@router.post("/admin/manual-results")
async def submit_manual_results(submission: ManualResultsSubmission):
    """
    Submit manual meet results
    - Creates meet if it doesn't exist
    - Validates athletes exist
    - Inserts results into database
    """
    conn = get_database_connection()
    cursor = conn.cursor()

    try:
        # 1. Create or get meet
        cursor.execute(
            "SELECT id FROM meets WHERE meet_name = ?",
            (submission.meet_name,)
        )
        meet_row = cursor.fetchone()

        if not meet_row:
            # Create new meet
            meet_id = f"manual_{datetime.now().timestamp()}"

            # Validate and convert meet_date to ISO 8601 format
            try:
                validated_meet_date = parse_and_validate_date(
                    submission.meet_date.strip(),
                    field_name="meet_date"
                )
            except ValueError as e:
                logger.error(f"Date validation failed for manual entry: {e}")
                raise HTTPException(status_code=400, detail=f"Invalid meet date format: {str(e)}")

            cursor.execute("""
                INSERT INTO meets (id, meet_name, meet_type, meet_date, meet_city, meet_course)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                meet_id,
                submission.meet_name.strip(),
                submission.meet_alias.strip() or "Manual Entry",
                validated_meet_date,
                submission.meetcity.strip(),
                submission.meet_course.strip()
            ))
            conn.commit()
        else:
            meet_id = meet_row[0]

        # 2. Validate and insert results
        inserted_count = 0
        errors = []

        for result in submission.results:
            try:
                # Check athlete exists
                cursor.execute(
                    "SELECT id FROM athletes WHERE id = ?",
                    (result.athlete_id,)
                )
                if not cursor.fetchone():
                    errors.append(f"Athlete {result.athlete_id} not found")
                    continue

                # Find or create event
                cursor.execute("""
                    SELECT id FROM events
                    WHERE event_distance = ? AND event_stroke = ? AND gender = ?
                """, (result.distance, result.stroke.upper(), result.event_gender.upper()))

                event_row = cursor.fetchone()
                if not event_row:
                    errors.append(f"Event {result.distance}{result.stroke} {result.event_gender} not found")
                    continue

                event_id = event_row[0]

                # Insert result
                cursor.execute("""
                    INSERT INTO results (
                        athlete_id, event_id, meet_id, time_string, comp_place,
                        club_name, state_code
                    ) VALUES (?, ?, ?, ?, ?, NULL, NULL)
                """, (
                    result.athlete_id,
                    event_id,
                    meet_id,
                    result.time_string.strip(),
                    result.comp_place
                ))
                inserted_count += 1

            except Exception as e:
                errors.append(f"Error inserting result for athlete {result.athlete_id}: {str(e)}")

        conn.commit()

        response = {
            "success": True,
            "meet_id": meet_id,
            "results_inserted": inserted_count,
            "total_results_submitted": len(submission.results),
            "message": f"Inserted {inserted_count}/{len(submission.results)} results for meet '{submission.meet_name}'"
        }

        if errors:
            response["errors"] = errors

        return response

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to submit results: {str(e)}")

    finally:
        conn.close()


@router.get("/admin/export-base-table/{table_type}")
async def export_base_table(table_type: str):
    """
    Export base time tables as Excel files.

    Args:
        table_type: 'map', 'mot', 'aqua', 'podium', or 'canada'

    Returns:
        Excel file download
    """
    import pandas as pd
    from io import BytesIO

    # Map table type to database table name
    table_map = {
        'map': 'map_base_times',
        'mot': 'mot_base_times',
        'aqua': 'aqua_base_times',
        'podium': 'podium_target_times',
        'canada': 'canada_on_track'
    }

    if table_type not in table_map:
        raise HTTPException(status_code=400, detail=f"Invalid table type. Must be one of: {list(table_map.keys())}")

    table_name = table_map[table_type]

    conn = get_database_connection()
    try:
        # Read the table into a DataFrame
        df = pd.read_sql_query(f"SELECT * FROM {table_name}", conn)

        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=table_type.upper())

        output.seek(0)

        # Return as downloadable file
        filename = f"{table_type}_base_times.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export {table_type} table: {str(e)}")

    finally:
        conn.close()


@router.get("/admin/podium-target-times")
async def get_podium_target_times(year: int = None):
    """
    Get podium target times, optionally filtered by year.
    Returns event details parsed from event_id for display.
    """
    conn = get_database_connection()
    try:
        cursor = conn.cursor()

        if year:
            cursor.execute("""
                SELECT id, event_id, target_time_seconds, sea_games_year
                FROM podium_target_times
                WHERE sea_games_year = ?
                ORDER BY event_id
            """, (year,))
        else:
            cursor.execute("""
                SELECT id, event_id, target_time_seconds, sea_games_year
                FROM podium_target_times
                ORDER BY sea_games_year DESC, event_id
            """)

        rows = cursor.fetchall()
        results = []
        for row in rows:
            # Parse event_id: LCM_Free_100_M -> course=LCM, stroke=Free, distance=100, gender=M
            event_id = row[1]
            parts = event_id.split('_')
            if len(parts) >= 4:
                course = parts[0]
                stroke = parts[1]
                distance = parts[2]
                gender = parts[3]
                event_display = f"{distance} {stroke}"
            else:
                event_display = event_id
                gender = "?"

            # Convert seconds to time string
            time_seconds = row[2]
            minutes = int(time_seconds // 60)
            secs = time_seconds % 60
            if minutes > 0:
                time_string = f"{minutes}:{secs:05.2f}"
            else:
                time_string = f"{secs:.2f}"

            results.append({
                "id": row[0],
                "event_id": event_id,
                "event_display": event_display,
                "gender": gender,
                "target_time_seconds": time_seconds,
                "target_time_string": time_string,
                "sea_games_year": row[3]
            })

        # Get distinct years for dropdown
        cursor.execute("SELECT DISTINCT sea_games_year FROM podium_target_times ORDER BY sea_games_year DESC")
        years = [r[0] for r in cursor.fetchall()]

        return {"times": results, "available_years": years}

    finally:
        conn.close()


@router.get("/admin/events-list")
async def get_events_list():
    """
    Get list of all individual LCM events for populating update forms.
    """
    conn = get_database_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, event_distance, event_stroke, gender
            FROM events
            WHERE event_course = 'LCM' AND event_type = 'Individual'
            ORDER BY gender,
                CASE event_stroke
                    WHEN 'Free' THEN 1
                    WHEN 'Back' THEN 2
                    WHEN 'Breast' THEN 3
                    WHEN 'Fly' THEN 4
                    WHEN 'Medley' THEN 5
                    ELSE 6
                END,
                event_distance
        """)

        rows = cursor.fetchall()
        events = []
        for row in rows:
            event_id = row[0]
            distance = row[1]
            stroke = row[2]
            gender = row[3]
            events.append({
                "event_id": event_id,
                "event_display": f"{distance} {stroke}",
                "gender": gender
            })

        return {"events": events}

    finally:
        conn.close()


@router.post("/admin/podium-target-times")
async def save_podium_target_times(request: Request):
    """
    Save/update podium target times for a specific year.
    Expects JSON body: { "year": 2025, "times": [{"event_id": "LCM_Free_100_M", "time_string": "49.69"}, ...] }
    """
    conn = get_database_connection()
    try:
        data = await request.json()
        year = data.get("year")
        times = data.get("times", [])

        if not year:
            raise HTTPException(status_code=400, detail="Year is required")

        cursor = conn.cursor()
        updated = 0
        inserted = 0

        for item in times:
            event_id = item.get("event_id")
            time_str = item.get("time_string", "").strip()

            if not event_id or not time_str:
                continue

            # Parse time string to seconds
            try:
                if ':' in time_str:
                    parts = time_str.split(':')
                    minutes = float(parts[0])
                    seconds = float(parts[1])
                    time_seconds = minutes * 60 + seconds
                else:
                    time_seconds = float(time_str)
            except ValueError:
                continue

            # Check if exists for this year
            cursor.execute("""
                SELECT id FROM podium_target_times
                WHERE event_id = ? AND sea_games_year = ?
            """, (event_id, year))

            existing = cursor.fetchone()
            if existing:
                cursor.execute("""
                    UPDATE podium_target_times
                    SET target_time_seconds = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (time_seconds, existing[0]))
                updated += 1
            else:
                cursor.execute("""
                    INSERT INTO podium_target_times (event_id, target_time_seconds, sea_games_year)
                    VALUES (?, ?, ?)
                """, (event_id, time_seconds, year))
                inserted += 1

        conn.commit()
        return {"success": True, "updated": updated, "inserted": inserted}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        conn.close()


@router.get("/admin/map-base-times")
async def get_map_base_times(age: int = None, year: int = None):
    """
    Get MAP base times, optionally filtered by age and/or year.
    Returns event details parsed from event_id for display.
    """
    conn = get_database_connection()
    try:
        cursor = conn.cursor()

        if age and year:
            cursor.execute("""
                SELECT id, event_id, age, base_time_seconds, competition_year
                FROM map_base_times
                WHERE age = ? AND competition_year = ?
                ORDER BY event_id
            """, (age, year))
        elif age:
            cursor.execute("""
                SELECT id, event_id, age, base_time_seconds, competition_year
                FROM map_base_times
                WHERE age = ?
                ORDER BY event_id
            """, (age,))
        elif year:
            cursor.execute("""
                SELECT id, event_id, age, base_time_seconds, competition_year
                FROM map_base_times
                WHERE competition_year = ?
                ORDER BY age, event_id
            """, (year,))
        else:
            cursor.execute("""
                SELECT id, event_id, age, base_time_seconds, competition_year
                FROM map_base_times
                ORDER BY competition_year DESC, age, event_id
            """)

        rows = cursor.fetchall()
        results = []
        for row in rows:
            event_id = row[1]
            parts = event_id.split('_') if event_id else []
            if len(parts) >= 4:
                stroke = parts[1]
                distance = parts[2]
                gender = parts[3]
                event_display = f"{distance} {stroke}"
            else:
                event_display = event_id or "Unknown"
                gender = "?"

            # Convert seconds to time string
            time_seconds = row[3]
            if time_seconds:
                minutes = int(time_seconds // 60)
                secs = time_seconds % 60
                if minutes > 0:
                    time_string = f"{minutes}:{secs:05.2f}"
                else:
                    time_string = f"{secs:.2f}"
            else:
                time_string = ""

            results.append({
                "id": row[0],
                "event_id": event_id,
                "event_display": event_display,
                "gender": gender,
                "age": int(row[2]),
                "base_time_seconds": time_seconds,
                "time_string": time_string,
                "competition_year": row[4] if len(row) > 4 else 2025
            })

        # Get distinct years for dropdown
        cursor.execute("SELECT DISTINCT competition_year FROM map_base_times ORDER BY competition_year DESC")
        available_years = [r[0] for r in cursor.fetchall() if r[0] is not None]

        return {"times": results, "available_years": available_years}

    finally:
        conn.close()


@router.post("/admin/map-base-times")
async def save_map_base_times(request: Request):
    """
    Save/update MAP base times for a specific age and year.
    Expects JSON body: { "age": 12, "year": 2025, "times": [{"event_id": "LCM_Free_100_M", "time_string": "49.69"}, ...] }
    """
    conn = get_database_connection()
    try:
        data = await request.json()
        age = data.get("age")
        year = data.get("year", 2025)  # Default to 2025 if not specified
        times = data.get("times", [])

        if not age:
            raise HTTPException(status_code=400, detail="Age is required")

        cursor = conn.cursor()
        updated = 0
        inserted = 0

        for item in times:
            event_id = item.get("event_id")
            time_str = item.get("time_string", "").strip()

            if not event_id or not time_str:
                continue

            # Parse time string to seconds
            try:
                if ':' in time_str:
                    parts = time_str.split(':')
                    minutes = float(parts[0])
                    seconds = float(parts[1])
                    time_seconds = minutes * 60 + seconds
                else:
                    time_seconds = float(time_str)
            except ValueError:
                continue

            # Parse gender from event_id
            event_parts = event_id.split('_')
            gender = event_parts[3] if len(event_parts) >= 4 else 'M'

            # Check if exists for this age, event, and year
            cursor.execute("""
                SELECT id FROM map_base_times
                WHERE event_id = ? AND age = ? AND competition_year = ?
            """, (event_id, age, year))

            existing = cursor.fetchone()
            if existing:
                cursor.execute("""
                    UPDATE map_base_times
                    SET base_time_seconds = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (time_seconds, existing[0]))
                updated += 1
            else:
                # Parse event details for legacy columns
                stroke = event_parts[1] if len(event_parts) >= 2 else ''
                distance = event_parts[2] if len(event_parts) >= 3 else ''
                event_name = f"{distance} {stroke}"

                cursor.execute("""
                    INSERT INTO map_base_times (event_id, gender, event, age, base_time_seconds, competition_year)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (event_id, gender, event_name, age, time_seconds, year))
                inserted += 1

        conn.commit()
        return {"success": True, "updated": updated, "inserted": inserted}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        conn.close()


@router.get("/admin/aqua-base-times")
async def get_aqua_base_times(year: int = None, course: str = None):
    """
    Get AQUA base times, optionally filtered by year and/or course (LCM/SCM).
    """
    conn = get_database_connection()
    try:
        cursor = conn.cursor()

        if year and course:
            cursor.execute("""
                SELECT id, event_id, base_time_seconds, course, competition_year
                FROM aqua_base_times
                WHERE competition_year = ? AND course = ?
                ORDER BY event_id
            """, (year, course))
        elif year:
            cursor.execute("""
                SELECT id, event_id, base_time_seconds, course, competition_year
                FROM aqua_base_times
                WHERE competition_year = ?
                ORDER BY course, event_id
            """, (year,))
        elif course:
            cursor.execute("""
                SELECT id, event_id, base_time_seconds, course, competition_year
                FROM aqua_base_times
                WHERE course = ?
                ORDER BY competition_year DESC, event_id
            """, (course,))
        else:
            cursor.execute("""
                SELECT id, event_id, base_time_seconds, course, competition_year
                FROM aqua_base_times
                ORDER BY competition_year DESC, course, event_id
            """)

        rows = cursor.fetchall()
        results = []
        for row in rows:
            event_id = row[1]
            parts = event_id.split('_') if event_id else []
            if len(parts) >= 4:
                stroke = parts[1]
                distance = parts[2]
                gender = parts[3]
                event_display = f"{distance} {stroke}"
            else:
                event_display = event_id or "Unknown"
                gender = "?"

            # Convert seconds to time string
            time_seconds = row[2]
            if time_seconds:
                minutes = int(time_seconds // 60)
                secs = time_seconds % 60
                if minutes > 0:
                    time_string = f"{minutes}:{secs:05.2f}"
                else:
                    time_string = f"{secs:.2f}"
            else:
                time_string = ""

            results.append({
                "id": row[0],
                "event_id": event_id,
                "event_display": event_display,
                "gender": gender,
                "base_time_seconds": time_seconds,
                "time_string": time_string,
                "course": row[3] or "LCM",
                "competition_year": row[4] or 2025
            })

        # Get distinct years and courses for dropdowns
        cursor.execute("SELECT DISTINCT competition_year FROM aqua_base_times WHERE competition_year IS NOT NULL ORDER BY competition_year DESC")
        available_years = [r[0] for r in cursor.fetchall()]

        cursor.execute("SELECT DISTINCT course FROM aqua_base_times WHERE course IS NOT NULL ORDER BY course")
        available_courses = [r[0] for r in cursor.fetchall()]

        return {"times": results, "available_years": available_years, "available_courses": available_courses}

    finally:
        conn.close()


@router.post("/admin/aqua-base-times")
async def save_aqua_base_times(request: Request):
    """
    Save/update AQUA base times for a specific year and course.
    Expects JSON body: { "year": 2025, "course": "LCM", "times": [{"event_id": "LCM_Free_100_M", "time_string": "46.40"}, ...] }
    """
    conn = get_database_connection()
    try:
        data = await request.json()
        year = data.get("year", 2025)
        course = data.get("course", "LCM")
        times = data.get("times", [])

        cursor = conn.cursor()
        updated = 0
        inserted = 0

        for item in times:
            event_id = item.get("event_id")
            time_str = item.get("time_string", "").strip()

            if not event_id or not time_str:
                continue

            # Parse time string to seconds
            try:
                if ':' in time_str:
                    parts = time_str.split(':')
                    minutes = float(parts[0])
                    seconds = float(parts[1])
                    time_seconds = minutes * 60 + seconds
                else:
                    time_seconds = float(time_str)
            except ValueError:
                continue

            # Parse event details from event_id
            event_parts = event_id.split('_')
            gender = event_parts[3] if len(event_parts) >= 4 else 'M'
            stroke = event_parts[1] if len(event_parts) >= 2 else ''
            distance = int(event_parts[2]) if len(event_parts) >= 3 else 0
            event_name = f"{distance} {stroke}"

            # Update event_id to use correct course prefix
            new_event_id = f"{course}_{stroke}_{distance}_{gender}"

            # Check if exists for this year, course, and event
            cursor.execute("""
                SELECT id FROM aqua_base_times
                WHERE gender = ? AND distance = ? AND stroke = ? AND course = ? AND competition_year = ?
            """, (gender, distance, stroke, course, year))

            existing = cursor.fetchone()
            if existing:
                cursor.execute("""
                    UPDATE aqua_base_times
                    SET base_time_seconds = ?, event_id = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (time_seconds, new_event_id, existing[0]))
                updated += 1
            else:
                cursor.execute("""
                    INSERT INTO aqua_base_times (event_id, gender, event, distance, stroke, base_time_seconds, course, competition_year)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (new_event_id, gender, event_name, distance, stroke, time_seconds, course, year))
                inserted += 1

        conn.commit()
        return {"success": True, "updated": updated, "inserted": inserted}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        conn.close()


@router.get("/admin/mot-base-times")
async def get_mot_base_times():
    """
    Get MOT base times for all events and ages.
    Returns event details parsed from mot_event_id for display.
    Ages 15-23 for 100m+ events, 18-23 for 50m events.
    """
    conn = get_database_connection()
    try:
        cursor = conn.cursor()

        cursor.execute("""
            SELECT id, mot_event_id, mot_age, mot_time_seconds
            FROM mot_base_times
            ORDER BY mot_event_id, mot_age
        """)

        rows = cursor.fetchall()
        results = []
        for row in rows:
            event_id = row[1]
            parts = event_id.split('_') if event_id else []
            if len(parts) >= 4:
                stroke = parts[1]
                distance = parts[2]
                gender = parts[3]
                event_display = f"{distance} {stroke}"
                is_50m = (distance == '50')
            else:
                event_display = event_id or "Unknown"
                gender = "?"
                is_50m = False

            # Convert seconds to time string
            time_seconds = row[3]
            if time_seconds:
                minutes = int(time_seconds // 60)
                secs = time_seconds % 60
                if minutes > 0:
                    time_string = f"{minutes}:{secs:05.2f}"
                else:
                    time_string = f"{secs:.2f}"
            else:
                time_string = ""

            results.append({
                "id": row[0],
                "event_id": event_id,
                "event_display": event_display,
                "gender": gender,
                "age": int(row[2]),
                "time_seconds": time_seconds,
                "time_string": time_string,
                "is_50m": is_50m
            })

        return {"times": results}

    finally:
        conn.close()


@router.post("/admin/mot-base-times")
async def save_mot_base_times(request: Request):
    """
    Save/update MOT base times for specific ages.
    Expects JSON body: { "times": [{"event_id": "LCM_Free_100_M", "age": 15, "time_string": "49.69"}, ...] }
    """
    conn = get_database_connection()
    try:
        data = await request.json()
        times = data.get("times", [])

        cursor = conn.cursor()
        updated = 0
        inserted = 0

        for item in times:
            event_id = item.get("event_id")
            age = item.get("age")
            time_str = item.get("time_string", "").strip()

            if not event_id or age is None or not time_str:
                continue

            # Parse time string to seconds
            try:
                if ':' in time_str:
                    parts = time_str.split(':')
                    minutes = float(parts[0])
                    seconds = float(parts[1])
                    time_seconds = minutes * 60 + seconds
                else:
                    time_seconds = float(time_str)
            except ValueError:
                continue

            # Check if exists for this event and age
            cursor.execute("""
                SELECT id FROM mot_base_times
                WHERE mot_event_id = ? AND mot_age = ?
            """, (event_id, age))

            existing = cursor.fetchone()
            if existing:
                cursor.execute("""
                    UPDATE mot_base_times
                    SET mot_time_seconds = ?
                    WHERE id = ?
                """, (time_seconds, existing[0]))
                updated += 1
            else:
                cursor.execute("""
                    INSERT INTO mot_base_times (mot_event_id, mot_age, mot_time_seconds)
                    VALUES (?, ?, ?)
                """, (event_id, age, time_seconds))
                inserted += 1

        conn.commit()
        return {"success": True, "updated": updated, "inserted": inserted}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        conn.close()


@router.get("/admin/meet-results/{meet_id}")
async def get_meet_results(meet_id: str):
    """
    Get all results for a specific meet with athlete names and event info for editing.
    """
    conn = get_database_connection()
    try:
        cursor = conn.cursor()

        # Get results with athlete name and event info
        cursor.execute("""
            SELECT
                r.id,
                r.athlete_id,
                r.foreign_athlete_id,
                COALESCE(a.fullname, fa.fullname, 'Unknown') as athlete_name,
                r.event_id,
                e.event_distance,
                e.event_stroke,
                r.time_string,
                r.time_seconds,
                r.comp_place,
                e.gender,
                r.result_status
            FROM results r
            LEFT JOIN athletes a ON r.athlete_id = a.id
            LEFT JOIN foreign_athletes fa ON r.foreign_athlete_id = fa.id
            LEFT JOIN events e ON r.event_id = e.id
            WHERE r.meet_id = ?
            ORDER BY
                CASE e.event_stroke
                    WHEN 'Free' THEN 1
                    WHEN 'Back' THEN 2
                    WHEN 'Breast' THEN 3
                    WHEN 'Fly' THEN 4
                    WHEN 'Medley' THEN 5
                    ELSE 6
                END,
                e.event_distance,
                e.gender,
                r.time_seconds
        """, (meet_id,))

        rows = cursor.fetchall()
        results = []
        for row in rows:
            distance = row[5]
            stroke = row[6]
            event_display = f"{distance} {stroke}" if distance and stroke else "Unknown"

            results.append({
                "id": row[0],
                "athlete_id": row[1],
                "foreign_athlete_id": row[2],
                "athlete_name": row[3],
                "event_id": row[4],
                "event_display": event_display,
                "time_string": row[7] or "",
                "time_seconds": row[8],
                "comp_place": row[9],
                "gender": row[10],
                "result_status": row[11] or "OK"
            })

        return {"results": results, "count": len(results)}

    finally:
        conn.close()


@router.post("/admin/meet-results/update-comp-place")
async def update_comp_place(request: Request):
    """
    Update comp_place and/or result_status for multiple result records.
    Expects JSON body: { "updates": [{"result_id": "uuid", "value": "1" or "DNS"}, ...] }

    If value is a number -> update comp_place, set result_status to OK
    If value is DQ/DNS/DNF/SCR -> update result_status, clear comp_place and time
    If value is empty -> clear comp_place, set result_status to OK
    """
    conn = get_database_connection()
    try:
        data = await request.json()
        updates = data.get("updates", [])

        cursor = conn.cursor()
        updated = 0
        status_values = {'DQ', 'DNS', 'DNF', 'SCR'}

        for item in updates:
            result_id = item.get("result_id")
            value = item.get("value", "")

            if result_id is None:
                continue

            # Check if value is a status code
            value_upper = str(value).strip().upper()

            if value_upper in status_values:
                # It's a status - update result_status, clear comp_place, time, and points
                cursor.execute("""
                    UPDATE results
                    SET result_status = ?, comp_place = NULL, time_string = NULL, time_seconds = NULL,
                        aqua_points = NULL, rudolph_points = NULL
                    WHERE id = ?
                """, (value_upper, result_id))
            elif value == "" or value is None:
                # Empty - clear comp_place, set status to OK
                cursor.execute("""
                    UPDATE results SET comp_place = NULL, result_status = 'OK' WHERE id = ?
                """, (result_id,))
            else:
                # It's a number - update comp_place, set status to OK
                try:
                    place_num = int(value)
                    cursor.execute("""
                        UPDATE results SET comp_place = ?, result_status = 'OK' WHERE id = ?
                    """, (place_num, result_id))
                except ValueError:
                    # Invalid value, skip
                    continue
            updated += 1

        conn.commit()
        return {"success": True, "updated": updated}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        conn.close()


@router.get("/admin/canada-on-track")
async def get_canada_on_track(year: int = None):
    """
    Get Canada On Track times, optionally filtered by year.
    Returns event details with track and time info.
    """
    conn = get_database_connection()
    try:
        cursor = conn.cursor()

        if year:
            cursor.execute("""
                SELECT id, event_id, canada_track, canada_track_age, canada_track_time_seconds, canada_track_year
                FROM canada_on_track
                WHERE canada_track_year = ?
                ORDER BY event_id, canada_track
            """, (year,))
        else:
            cursor.execute("""
                SELECT id, event_id, canada_track, canada_track_age, canada_track_time_seconds, canada_track_year
                FROM canada_on_track
                ORDER BY canada_track_year DESC, event_id, canada_track
            """)

        rows = cursor.fetchall()
        results = []
        for row in rows:
            event_id = row[1]
            time_seconds = row[4]

            # Convert seconds to time string
            if time_seconds >= 60:
                minutes = int(time_seconds // 60)
                secs = time_seconds % 60
                time_string = f"{minutes}:{secs:05.2f}"
            else:
                time_string = f"{time_seconds:.2f}"

            results.append({
                "id": row[0],
                "event_id": event_id,
                "canada_track": row[2],
                "canada_track_age": row[3],
                "time_string": time_string,
                "time_seconds": time_seconds,
                "canada_track_year": row[5]
            })

        # Get available years
        cursor.execute("SELECT DISTINCT canada_track_year FROM canada_on_track ORDER BY canada_track_year DESC")
        available_years = [r[0] for r in cursor.fetchall()]

        return {
            "times": results,
            "available_years": available_years
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        conn.close()


@router.post("/admin/canada-on-track")
async def save_canada_on_track(request: Request):
    """
    Save Canada On Track times.
    Expects: { year: number, times: [{ event_id, track, age, time_string }] }
    """
    conn = get_database_connection()
    try:
        data = await request.json()
        year = data.get('year', 2025)
        times = data.get('times', [])

        cursor = conn.cursor()
        updated = 0
        inserted = 0

        for t in times:
            event_id = t.get('event_id')
            track = t.get('track')
            age = t.get('age')
            time_string = t.get('time_string', '').strip()

            if not event_id or not time_string or track is None or age is None:
                continue

            # Parse time string to seconds
            time_seconds = None
            if ':' in time_string:
                parts = time_string.split(':')
                if len(parts) == 2:
                    minutes = int(parts[0])
                    secs = float(parts[1])
                    time_seconds = minutes * 60 + secs
            else:
                time_seconds = float(time_string)

            if time_seconds is None:
                continue

            # Check if exists for this event_id, track, age, and year
            cursor.execute("""
                SELECT id FROM canada_on_track
                WHERE event_id = ? AND canada_track = ? AND canada_track_age = ? AND canada_track_year = ?
            """, (event_id, track, age, year))
            existing = cursor.fetchone()

            if existing:
                cursor.execute("""
                    UPDATE canada_on_track
                    SET canada_track_time_seconds = ?
                    WHERE id = ?
                """, (time_seconds, existing[0]))
                updated += 1
            else:
                cursor.execute("""
                    INSERT INTO canada_on_track (event_id, canada_track, canada_track_age, canada_track_time_seconds, canada_track_year)
                    VALUES (?, ?, ?, ?, ?)
                """, (event_id, track, age, time_seconds, year))
                inserted += 1

        conn.commit()
        return {"success": True, "updated": updated, "inserted": inserted}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        conn.close()
